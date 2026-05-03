from __future__ import annotations

import json
import logging
import time as pytime
import xml.etree.ElementTree as ET
import csv
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, time, timedelta
from datetime import timezone as dt_timezone
from io import BytesIO, StringIO
from threading import Lock, Thread

from django.conf import settings
from django.core.cache import cache
from django.db import close_old_connections
from django.db.models import Q
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from django.views.decorators.http import require_GET
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from rest_framework.decorators import api_view
from rest_framework.decorators import permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from hik_gateway.client import HikGatewayClient  # backward-compatible import for tests
from hik_gateway.models import AttendanceCorrection, AttendanceCorrectionLog, AttendanceLog, Device
from hik_gateway.services.device_payload import extract_devices, normalize_device
from hik_gateway.services.device_dispatch import dispatch_hik_devices_to_core_devices
from hik_gateway.services.device_sync import sync_all_gateways
from hik_gateway.services.gateway_connection import get_shared_gateway_client
from hik_gateway.services.catchup import (
    catchup_all_devices,
    catchup_tenant_devices_fast,
)
from hik_gateway.services.webhook_ingest import ingest_event
from employees.models import Employee
from employees.schedule_resolver import ScheduleResolver
from tenants.models import Tenant, TenantRole
from tenants.services import has_tenant_role


logger = logging.getLogger(__name__)
_AUTO_CATCHUP_LOCK = Lock()
_AUTO_CATCHUP_IN_FLIGHT: set[str] = set()


DEFAULT_DEVICE_LIST_PAYLOAD = {
    "SearchDescription": {
        "position": 0,
        "maxResult": 100,
        "Filter": {
            "key": "",
            "devType": "",
            "protocolType": ["ehomeV5"],
            "devStatus": ["online", "offline"],
        },
    }
}


def _safe_json_preview(payload: object, *, limit: int = 800) -> str:
    try:
        text = json.dumps(payload, ensure_ascii=False, default=str)
    except TypeError:
        text = str(payload)
    if len(text) <= limit:
        return text
    return f"{text[:limit]}...(+{len(text) - limit} chars)"


def _acs_event_summary(payload: dict | None) -> dict:
    events = _extract_acs_events(payload)
    with_card = 0
    serials: list[int] = []
    times: list[str] = []
    for event in events:
        card_no = str(event.get("cardNo") or event.get("card_no") or event.get("cardNumber") or "").strip()
        if card_no:
            with_card += 1
        raw_serial = event.get("serialNo") or event.get("serial_no")
        try:
            if raw_serial is not None:
                serials.append(int(raw_serial))
        except (TypeError, ValueError):
            pass
        time_value = str(event.get("dateTime") or event.get("time") or event.get("eventTime") or "").strip()
        if time_value:
            times.append(time_value)
    first_event = events[0] if events else {}
    return {
        "events_total": len(events),
        "events_with_card": with_card,
        "serial_min": min(serials) if serials else None,
        "serial_max": max(serials) if serials else None,
        "time_min": min(times) if times else None,
        "time_max": max(times) if times else None,
        "first_event_keys": sorted(first_event.keys())[:20] if isinstance(first_event, dict) else [],
        "first_event_preview": {
            "cardNo": first_event.get("cardNo") if isinstance(first_event, dict) else None,
            "card_no": first_event.get("card_no") if isinstance(first_event, dict) else None,
            "cardNumber": first_event.get("cardNumber") if isinstance(first_event, dict) else None,
            "dateTime": first_event.get("dateTime") if isinstance(first_event, dict) else None,
            "time": first_event.get("time") if isinstance(first_event, dict) else None,
            "eventTime": first_event.get("eventTime") if isinstance(first_event, dict) else None,
            "serialNo": first_event.get("serialNo") if isinstance(first_event, dict) else None,
            "serial_no": first_event.get("serial_no") if isinstance(first_event, dict) else None,
            "major": first_event.get("major") if isinstance(first_event, dict) else None,
            "minor": first_event.get("minor") if isinstance(first_event, dict) else None,
        },
    }


def _client_ip(request: HttpRequest) -> str:
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "")


def _is_allowed_ip(ip: str) -> bool:
    allowed = getattr(settings, "HIK_GATEWAY_ALLOWED_IPS", [])
    if not allowed:
        return True
    return ip in allowed


def _resolve_tenant(request: HttpRequest, payload: dict) -> Tenant | None:
    tenant_code = request.headers.get("X-TENANT-CODE", "").strip()
    root = payload.get("EventNotificationAlert", payload) if isinstance(payload, dict) else {}
    if not tenant_code and isinstance(root, dict):
        tenant_code = str(root.get("tenantCode") or payload.get("tenantCode") or "").strip()

    if not tenant_code:
        return None

    return Tenant.objects.filter(code=tenant_code).first()


def _is_allowed_token(request: HttpRequest) -> bool:
    expected = getattr(settings, "HIK_GATEWAY_WEBHOOK_TOKEN", "")
    if not expected:
        return True
    provided = request.headers.get("X-HIK-TOKEN", "")
    return provided == expected


def _strip_xml_ns(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


def _xml_element_to_dict(element: ET.Element) -> dict:
    payload: dict = {}
    for child in element:
        key = _strip_xml_ns(child.tag)
        if list(child):
            value = _xml_element_to_dict(child)
        else:
            value = (child.text or "").strip()
        if key in payload:
            if not isinstance(payload[key], list):
                payload[key] = [payload[key]]
            payload[key].append(value)
        else:
            payload[key] = value
    return payload


def _parse_webhook_payload(raw_body: str) -> dict | None:
    try:
        parsed = json.loads(raw_body or "{}")
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    # Some Hikvision firmwares prepend plain-text header lines before JSON.
    json_start = raw_body.find("{")
    json_end = raw_body.rfind("}")
    if json_start != -1 and json_end > json_start:
        candidate = raw_body[json_start:json_end + 1]
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    try:
        root = ET.fromstring(raw_body or "")
    except ET.ParseError:
        return None

    root_name = _strip_xml_ns(root.tag)
    return {root_name: _xml_element_to_dict(root)}


@csrf_exempt
@require_POST
def hik_event_webhook(request: HttpRequest) -> JsonResponse:
    ip = _client_ip(request)
    if not _is_allowed_ip(ip) or not _is_allowed_token(request):
        return JsonResponse({"detail": "Unauthorized source"}, status=403)

    raw_body = request.body.decode("utf-8", errors="replace")
    logger.info("Hikvision webhook payload received", extra={"client_ip": ip, "raw_body": raw_body})

    payload = _parse_webhook_payload(raw_body)
    if payload is None:
        logger.warning("Unsupported webhook payload format from %s: %s", ip, raw_body[:500])
        return JsonResponse({"detail": "Invalid payload format", "preview": raw_body[:200]}, status=400)

    tenant = _resolve_tenant(request, payload)
    if request.headers.get("X-TENANT-CODE") and tenant is None:
        return JsonResponse({"detail": "Unknown tenant"}, status=400)

    raw_event, attendance = ingest_event(payload, source=AttendanceLog.SOURCE_REALTIME, tenant=tenant)
    if raw_event is None:
        if getattr(settings, "DEBUG", False):
            sample_key = "hik:webhook:ignored:sample"
            if cache.add(sample_key, 1, timeout=20):
                root = payload.get("EventNotificationAlert", payload) if isinstance(payload, dict) else {}
                nested = (
                    root.get("AccessControllerEvent")
                    or root.get("AcsEvent")
                    or root.get("AcsEventInfo")
                    or {}
                )
                logger.warning(
                    "Ignored Hikvision webhook payload event_type=%s dev_index=%s keys=%s nested_keys=%s preview=%s",
                    root.get("eventType") if isinstance(root, dict) else None,
                    root.get("devIndex") if isinstance(root, dict) else None,
                    sorted(root.keys())[:20] if isinstance(root, dict) else [],
                    sorted(nested.keys())[:20] if isinstance(nested, dict) else [],
                    _safe_json_preview(root),
                )
        return JsonResponse({"status": "ignored"}, status=202)

    return JsonResponse(
        {
            "status": "ok",
            "raw_event_id": raw_event.id,
            "attendance_log_id": attendance.id if attendance else None,
        },
        status=201,
    )


def _parse_csv_query_list(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def _to_bool(value: str | None) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _start_tenant_auto_catchup(*, tenant_code: str, max_results: int = 100) -> bool:
    """Run tenant catchup in a daemon thread so polling responses stay fast."""

    tenant_key = tenant_code.lower()
    with _AUTO_CATCHUP_LOCK:
        if tenant_key in _AUTO_CATCHUP_IN_FLIGHT:
            return False
        _AUTO_CATCHUP_IN_FLIGHT.add(tenant_key)

    def _run() -> None:
        close_old_connections()
        try:
            catchup_tenant_devices_fast(tenant_code=tenant_code, max_results=max_results)
        except Exception:  # noqa: BLE001
            logger.exception("Automatic catchup failed", extra={"tenant_code": tenant_code})
        finally:
            close_old_connections()
            with _AUTO_CATCHUP_LOCK:
                _AUTO_CATCHUP_IN_FLIGHT.discard(tenant_key)

    Thread(
        target=_run,
        name=f"hik-auto-catchup-{tenant_key}",
        daemon=True,
    ).start()
    return True


def _is_admin_request(request: HttpRequest) -> bool:
    user = getattr(request, "user", None)
    return bool(user and user.is_authenticated and (user.is_staff or user.is_superuser))


def _is_authenticated_request(request: HttpRequest) -> bool:
    user = getattr(request, "user", None)
    return bool(user and user.is_authenticated)


def _require_admin_api(request: HttpRequest) -> Response | None:
    if _is_admin_request(request):
        return None
    return Response(
        {"detail": "Admin privileges required."},
        status=status.HTTP_403_FORBIDDEN,
    )


def _resolve_tenant_from_code(tenant_code: str) -> Tenant | None:
    code = str(tenant_code or "").strip()
    if not code:
        return None
    return Tenant.objects.filter(code__iexact=code).first()


def _require_tenant_scope_api(request: HttpRequest, tenant_code: str) -> tuple[Response | None, Tenant | None]:
    tenant = _resolve_tenant_from_code(tenant_code)
    if tenant is None:
        return Response({"detail": "Unknown tenant."}, status=status.HTTP_404_NOT_FOUND), None
    if _is_admin_request(request):
        return None, tenant
    if has_tenant_role(request.user, tenant, TenantRole.VIEWER):
        return None, tenant
    return Response({"detail": "Insufficient tenant scope for this tenant."}, status=status.HTTP_403_FORBIDDEN), None


def _normalize_acs_event_cond(payload: dict, *, default_max_results: int) -> dict:
    def _to_int_or_value(value, default=None):
        if value is None:
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return value

    if "AcsEventCond" in payload and isinstance(payload["AcsEventCond"], dict):
        cond = dict(payload["AcsEventCond"])
    else:
        cond = {
            "searchID": payload.get("searchID", ""),
            "searchResultPosition": _to_int_or_value(payload.get("searchResultPosition", 0), 0),
            "maxResults": _to_int_or_value(payload.get("maxResults", default_max_results), default_max_results),
        }
        if payload.get("startTime"):
            cond["startTime"] = payload.get("startTime")
        if payload.get("endTime"):
            cond["endTime"] = payload.get("endTime")

    if "searchResultPosition" in cond:
        cond["searchResultPosition"] = _to_int_or_value(cond.get("searchResultPosition"), 0)
    if "maxResults" in cond:
        cond["maxResults"] = _to_int_or_value(cond.get("maxResults"), default_max_results)
    return {"AcsEventCond": cond}


def _acs_event_payload_without_time_window(payload: dict) -> dict | None:
    if not isinstance(payload, dict):
        return None
    cond = payload.get("AcsEventCond")
    if not isinstance(cond, dict):
        return None
    if "startTime" not in cond and "endTime" not in cond:
        return None
    fallback_cond = dict(cond)
    fallback_cond.pop("startTime", None)
    fallback_cond.pop("endTime", None)
    return {"AcsEventCond": fallback_cond}


def _extract_acs_events(payload: dict | None) -> list[dict]:
    if not isinstance(payload, dict):
        return []

    info = payload.get("AcsEventTotalNum") or payload.get("AcsEvent") or payload
    if not isinstance(info, dict):
        return []

    events = info.get("InfoList", [])
    if isinstance(events, dict):
        events = events.get("AcsEventInfo", [])
    elif isinstance(events, tuple):
        events = list(events)

    if not isinstance(events, list):
        return []

    normalized: list[dict] = []
    for event in events:
        if not isinstance(event, dict):
            continue
        nested = event.get("AcsEventInfo")
        if isinstance(nested, dict):
            normalized.append(nested)
        else:
            normalized.append(event)
    return normalized


def _extract_max_event_serial(payload: dict | None) -> int | None:
    max_serial: int | None = None
    for event in _extract_acs_events(payload):
        raw_serial = event.get("serialNo") or event.get("serial_no")
        try:
            serial_no = int(raw_serial)
        except (TypeError, ValueError):
            continue
        if max_serial is None or serial_no > max_serial:
            max_serial = serial_no
    return max_serial


def _parse_gateway_event_time(value: str | None) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None

    parsed = parse_datetime(text)
    if parsed is None:
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return None

    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, dt_timezone.utc)
    return parsed.astimezone(dt_timezone.utc)


def _extract_latest_card_event(
    gateway_payload: dict | None,
    *,
    started_at: datetime,
    tolerance_seconds: int = 5,
    min_serial_exclusive: int | None = None,
    diagnostics: dict | None = None,
) -> dict | None:
    candidates: list[tuple[datetime, int, dict, str]] = []
    threshold = started_at - timedelta(seconds=max(0, tolerance_seconds))
    scanned = 0
    rejected_missing_card = 0
    rejected_too_old = 0
    rejected_not_newer_serial = 0

    for event in _extract_acs_events(gateway_payload):
        scanned += 1
        card_no = str(event.get("cardNo") or event.get("card_no") or "").strip()
        if not card_no:
            rejected_missing_card += 1
            continue

        raw_serial = event.get("serialNo") or event.get("serial_no") or 0
        try:
            serial_no = int(raw_serial)
        except (TypeError, ValueError):
            serial_no = 0

        event_time = _parse_gateway_event_time(event.get("dateTime") or event.get("time") or event.get("eventTime"))
        is_newer_than_baseline = min_serial_exclusive is not None and serial_no > min_serial_exclusive

        is_recent_by_time = bool(event_time and event_time >= threshold)
        if min_serial_exclusive is not None and not is_recent_by_time and not is_newer_than_baseline:
            rejected_not_newer_serial += 1
            continue
        if event_time and event_time < threshold and not is_newer_than_baseline:
            rejected_too_old += 1
            continue

        candidates.append((event_time or started_at, serial_no, event, card_no))

    if not candidates:
        if diagnostics is not None:
            diagnostics.update(
                {
                    "scanned": scanned,
                    "candidates": 0,
                    "rejected_missing_card": rejected_missing_card,
                    "rejected_too_old": rejected_too_old,
                    "rejected_not_newer_serial": rejected_not_newer_serial,
                    "threshold": threshold.isoformat(),
                    "min_serial_exclusive": min_serial_exclusive,
                }
            )
        return None

    candidates.sort(key=lambda item: (item[0], item[1]))
    event_time, serial_no, event, card_no = candidates[-1]
    if diagnostics is not None:
        diagnostics.update(
            {
                "scanned": scanned,
                "candidates": len(candidates),
                "rejected_missing_card": rejected_missing_card,
                "rejected_too_old": rejected_too_old,
                "rejected_not_newer_serial": rejected_not_newer_serial,
                "threshold": threshold.isoformat(),
                "min_serial_exclusive": min_serial_exclusive,
                "selected_card_no": card_no,
                "selected_event_time": event_time.isoformat(),
                "selected_serial_no": serial_no or None,
            }
        )

    return {
        "card_no": card_no,
        "event_time": event_time.isoformat(),
        "serial_no": serial_no or None,
        "card_reader_no": event.get("cardReaderNo"),
        "door_no": event.get("doorNo"),
        "employee_no": str(event.get("employeeNo") or "").strip(),
        "employee_no_string": str(event.get("employeeNoString") or "").strip(),
    }


def _is_bad_json_error_message(exc: Exception) -> bool:
    message = str(exc).lower()
    return (
        "badjsoncontent" in message
        or "wrong json content" in message
        or "bad json content" in message
    )


def _extract_acs_info_block(payload: dict | None) -> dict:
    if not isinstance(payload, dict):
        return {}
    info = payload.get("AcsEventTotalNum") or payload.get("AcsEvent") or payload
    return info if isinstance(info, dict) else {}


def _to_int_or_none(value) -> int | None:
    try:
        if value is None:
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _build_tail_acs_event_payload(request_payload: dict, gateway_response: dict | None) -> dict | None:
    cond = request_payload.get("AcsEventCond")
    if not isinstance(cond, dict):
        return None

    info = _extract_acs_info_block(gateway_response)
    total_matches = _to_int_or_none(info.get("totalMatches"))
    max_results = _to_int_or_none(cond.get("maxResults")) or 30
    current_position = _to_int_or_none(cond.get("searchResultPosition")) or 0
    if total_matches is None or total_matches <= 0 or max_results <= 0:
        return None

    tail_position = max(total_matches - max_results, 0)
    if tail_position <= current_position:
        return None

    tail_cond = dict(cond)
    tail_cond["searchResultPosition"] = tail_position
    return {"AcsEventCond": tail_cond}


def _parse_iso_date(value: str | None, field_name: str) -> date | None:
    if not str(value or "").strip():
        return None
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        raise ValueError(f"{field_name} must use YYYY-MM-DD format")


def _parse_iso_time(value: str | None, field_name: str, *, required: bool = False) -> time | None:
    text = str(value or "").strip()
    if not text:
        if required:
            raise ValueError(f"{field_name} is required")
        return None
    try:
        return time.fromisoformat(text)
    except ValueError:
        raise ValueError(f"{field_name} must use HH:MM or HH:MM:SS format")


def _parse_optional_decimal(value, field_name: str) -> Decimal | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name} must be a decimal number")
    if parsed < 0:
        raise ValueError(f"{field_name} must be greater than or equal to 0")
    return parsed


def _serialize_attendance_correction(correction: AttendanceCorrection) -> dict:
    return {
        "id": correction.id,
        "tenant": correction.tenant.code,
        "person_id": correction.employee.employee_no,
        "employee_name": correction.employee.full_name or correction.employee.employee_no,
        "date": correction.work_date.isoformat(),
        "arrival_time": (
            correction.arrival_time.isoformat(timespec="minutes")
            if correction.arrival_time is not None
            else None
        ),
        "departure_time": (
            correction.departure_time.isoformat(timespec="minutes")
            if correction.departure_time is not None
            else None
        ),
        "break_start_time": (
            correction.break_start_time.isoformat(timespec="minutes")
            if correction.break_start_time is not None
            else None
        ),
        "break_end_time": (
            correction.break_end_time.isoformat(timespec="minutes")
            if correction.break_end_time is not None
            else None
        ),
        "overtime_hours": float(correction.overtime_hours) if correction.overtime_hours is not None else None,
        "notes": correction.notes or "",
        "created_at": correction.created_at.isoformat() if correction.created_at is not None else None,
        "updated_at": correction.updated_at.isoformat() if correction.updated_at is not None else None,
        "created_by": correction.created_by.username if correction.created_by is not None else None,
        "updated_by": correction.updated_by.username if correction.updated_by is not None else None,
    }


def _serialize_attendance_correction_log(log_entry: AttendanceCorrectionLog) -> dict:
    payload = log_entry.payload or {}
    reason = ""
    if isinstance(payload, dict):
        after_payload = payload.get("after")
        before_payload = payload.get("before")
        if isinstance(after_payload, dict):
            reason = str(after_payload.get("notes") or "").strip()
        if not reason and isinstance(before_payload, dict):
            reason = str(before_payload.get("notes") or "").strip()

    return {
        "id": log_entry.id,
        "tenant": log_entry.tenant.code,
        "person_id": log_entry.employee.employee_no,
        "employee_name": log_entry.employee.full_name or log_entry.employee.employee_no,
        "date": log_entry.work_date.isoformat(),
        "action": log_entry.action,
        "payload": log_entry.payload or {},
        "reason": reason,
        "changed_by": log_entry.changed_by.username if log_entry.changed_by is not None else None,
        "created_at": log_entry.created_at.isoformat() if log_entry.created_at is not None else None,
    }


def _resolve_report_window(
    period: str,
    *,
    start_date_value: str | None,
    end_date_value: str | None,
    reference_date_value: str | None,
) -> tuple[date, date]:
    start_date = _parse_iso_date(start_date_value, "start_date")
    end_date = _parse_iso_date(end_date_value, "end_date")
    reference_date = _parse_iso_date(reference_date_value, "date") or timezone.localdate()

    if start_date and end_date:
        if end_date < start_date:
            raise ValueError("end_date must be greater than or equal to start_date")
        return start_date, end_date

    anchor_date = start_date or end_date or reference_date
    if period == "daily":
        return anchor_date, anchor_date
    if period == "weekly":
        week_start = anchor_date - timedelta(days=anchor_date.weekday())
        return week_start, week_start + timedelta(days=6)

    month_start = anchor_date.replace(day=1)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    return month_start, next_month - timedelta(days=1)


def _classify_attendance_direction(log: AttendanceLog) -> str:
    action = str(log.normalized_action or "").strip().upper()
    if action in {AttendanceLog.ACTION_CHECK_IN, AttendanceLog.ACTION_OVERTIME_IN}:
        return "IN"
    if action in {AttendanceLog.ACTION_CHECK_OUT, AttendanceLog.ACTION_OVERTIME_OUT}:
        return "OUT"
    if action in {AttendanceLog.ACTION_BREAK_IN, AttendanceLog.ACTION_BREAK_OUT, AttendanceLog.ACTION_ACCESS_DENIED}:
        return "UNKNOWN"

    direction = str(log.direction or "").strip().upper()
    if direction in {"IN", "OUT"}:
        return direction

    combined = " ".join(
        [
            str(log.attendance_type or "").strip().lower(),
            str(log.attendance_status or "").strip().lower(),
        ]
    )
    if any(token in combined for token in ("checkin", "in", "entry", "entree", "arrival")):
        return "IN"
    if any(token in combined for token in ("checkout", "out", "exit", "sortie", "departure")):
        return "OUT"
    return "UNKNOWN"


def _iterate_dates(start_date: date, end_date: date):
    cursor = start_date
    while cursor <= end_date:
        yield cursor
        cursor += timedelta(days=1)


def _expected_bounds_from_day_schedule(
    day_schedule: dict,
    *,
    target_day: date,
    current_tz,
) -> tuple[datetime | None, datetime | None]:
    slots = day_schedule.get("slots") or []
    window_start = None
    window_end = None
    for slot in slots:
        if slot.get("slot_type") == "rest":
            continue
        start_time = slot.get("start_time")
        end_time = slot.get("end_time")
        if start_time is None or end_time is None:
            continue
        start_dt = timezone.make_aware(datetime.combine(target_day, start_time), current_tz)
        end_day = target_day if end_time > start_time else (target_day + timedelta(days=1))
        end_dt = timezone.make_aware(datetime.combine(end_day, end_time), current_tz)
        if window_start is None or start_dt < window_start:
            window_start = start_dt
        if window_end is None or end_dt > window_end:
            window_end = end_dt
    return window_start, window_end


def _minutes_delta(actual_value: datetime | None, expected_value: datetime | None) -> int | None:
    if actual_value is None or expected_value is None:
        return None
    return int((actual_value - expected_value).total_seconds() // 60)


def _minutes_to_hours(minutes: int | None) -> float:
    if minutes is None:
        return 0.0
    return round(max(minutes, 0) / 60, 2)


def _covers_at_least_one_shift(
    *,
    day_schedule: dict,
    target_day: date,
    current_tz,
    actual_checkin_at: datetime | None,
    actual_checkout_at: datetime | None,
) -> bool:
    if actual_checkin_at is None or actual_checkout_at is None:
        return False

    slots = day_schedule.get("slots") or []
    for slot in slots:
        if slot.get("slot_type") == "rest":
            continue
        start_time = slot.get("start_time")
        end_time = slot.get("end_time")
        if start_time is None or end_time is None:
            continue

        shift_start = timezone.make_aware(datetime.combine(target_day, start_time), current_tz)
        shift_end_day = target_day if end_time > start_time else (target_day + timedelta(days=1))
        shift_end = timezone.make_aware(datetime.combine(shift_end_day, end_time), current_tz)
        if actual_checkin_at <= shift_start and actual_checkout_at >= shift_end:
            return True
    return False


def _is_overtime_enabled_for_day(
    *,
    resolver: ScheduleResolver,
    employee: Employee,
    target_day: date,
) -> bool:
    resolution = resolver.resolve_assignment(employee, target_day)
    flags = []
    for key in ("planning", "work_shift"):
        resolved = resolution.get(key) if isinstance(resolution, dict) else None
        assignment = getattr(resolved, "assignment", None)
        if assignment is not None:
            flags.append(bool(getattr(assignment, "effective_for_overtime", True)))
    if not flags:
        return True
    return all(flags)


def _calculate_worked_minutes(
    actual_checkin_at: datetime | None,
    actual_checkout_at: datetime | None,
) -> int:
    if actual_checkin_at is None or actual_checkout_at is None:
        return 0
    delta_minutes = int((actual_checkout_at - actual_checkin_at).total_seconds() // 60)
    return max(delta_minutes, 0)


def _extract_validation_status(payload: dict | None) -> str:
    if not isinstance(payload, dict):
        return ""
    after_payload = payload.get("after")
    if not isinstance(after_payload, dict):
        after_payload = {}
    for key in (
        "validation_status",
        "rh_validation_status",
        "manager_validation_status",
    ):
        value = str(payload.get(key) or after_payload.get(key) or "").strip().lower()
        if value:
            return value
    return ""


def _build_detail_anomalies(
    *,
    status_key: str,
    has_work_period: bool,
    has_checkin: bool,
    has_checkout: bool,
    arrival_delta_minutes: int | None,
    departure_delta_minutes: int | None,
    unknown_events: int,
) -> list[str]:
    anomalies = []
    if has_work_period and not has_checkin and not has_checkout:
        anomalies.append("absence")
    if has_work_period and (has_checkin ^ has_checkout):
        anomalies.append("pointage_incomplet")
    if status_key == "unexpected_activity":
        anomalies.append("activite_inattendue")
    if arrival_delta_minutes is not None and arrival_delta_minutes > 0:
        anomalies.append("retard")
    if departure_delta_minutes is not None and departure_delta_minutes < 0:
        anomalies.append("depart_anticipe")
    if unknown_events > 0:
        anomalies.append("anomalie_inconnue")
    return anomalies


def _format_export_datetime(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return timezone.localtime(value).strftime("%Y-%m-%d %H:%M")
    return str(value)


def _format_export_time(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return timezone.localtime(value).strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value)


def _build_export_filename(prefix: str, start_date: date, end_date: date, extension: str) -> str:
    return f"{prefix}-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.{extension}"


ATTENDANCE_EXPORT_FIELD_CONFIG: dict[str, dict[str, str]] = {
    "tenant": {"header": "Tenant", "row_key": "tenant"},
    "person_id": {"header": "Person ID", "row_key": "person_id"},
    "employee_name": {"header": "Employe", "row_key": "employee_name"},
    "department_name": {"header": "Departement", "row_key": "department_name"},
    "planning_name": {"header": "Planning", "row_key": "planning_name"},
    "work_shift_name": {"header": "Shift", "row_key": "work_shift_name"},
    "date": {"header": "Date", "row_key": "date"},
    "status": {"header": "Statut", "row_key": "status"},
    "expected_work_period": {"header": "Pointage attendu", "row_key": "expected_work_period"},
    "planned_minutes": {"header": "Minutes planifiees", "row_key": "planned_minutes"},
    "total_logs": {"header": "Total logs", "row_key": "total_logs"},
    "checkins": {"header": "Entrees", "row_key": "checkins"},
    "checkouts": {"header": "Sorties", "row_key": "checkouts"},
    "unknown_events": {"header": "Inconnus", "row_key": "unknown_events"},
    "arrival_time": {"header": "Heure arrivee", "row_key": "arrival_time"},
    "departure_time": {"header": "Heure depart", "row_key": "departure_time"},
    "expected_checkin_at": {"header": "Arrivee attendue", "row_key": "expected_checkin_at"},
    "actual_checkin_at": {"header": "Arrivee reelle", "row_key": "actual_checkin_at"},
    "arrival_delta_minutes": {"header": "Ecart arrivee (min)", "row_key": "arrival_delta_minutes"},
    "expected_checkout_at": {"header": "Depart attendu", "row_key": "expected_checkout_at"},
    "actual_checkout_at": {"header": "Depart reel", "row_key": "actual_checkout_at"},
    "departure_delta_minutes": {"header": "Ecart depart (min)", "row_key": "departure_delta_minutes"},
}

ATTENDANCE_EXPORT_DEFAULT_FIELD_IDS = [
    "person_id",
    "employee_name",
    "department_name",
    "date",
    "arrival_time",
    "departure_time",
    "status",
]


def _resolve_attendance_export_fields(requested_fields: list[str]) -> list[str]:
    selected: list[str] = []
    for field in requested_fields:
        normalized = str(field or "").strip().lower()
        if not normalized or normalized in selected:
            continue
        if normalized in ATTENDANCE_EXPORT_FIELD_CONFIG:
            selected.append(normalized)
    if selected:
        return selected
    return ATTENDANCE_EXPORT_DEFAULT_FIELD_IDS.copy()


def _build_attendance_export_matrix(detail_rows: list[dict], selected_field_ids: list[str]) -> tuple[list[str], list[list]]:
    resolved_field_ids = [field_id for field_id in selected_field_ids if field_id in ATTENDANCE_EXPORT_FIELD_CONFIG]
    if not resolved_field_ids:
        resolved_field_ids = ATTENDANCE_EXPORT_DEFAULT_FIELD_IDS.copy()

    headers = [ATTENDANCE_EXPORT_FIELD_CONFIG[field_id]["header"] for field_id in resolved_field_ids]
    matrix_rows: list[list] = []
    for row in detail_rows:
        matrix_rows.append(
            [
                row.get(ATTENDANCE_EXPORT_FIELD_CONFIG[field_id]["row_key"], "")
                for field_id in resolved_field_ids
            ]
        )
    return headers, matrix_rows


def _build_attendance_export_rows(compliance_employees: list[dict]) -> list[dict]:
    status_labels = {
        "compliant": "Conforme",
        "partial": "Partiel",
        "missing": "Manquant",
        "unexpected_activity": "Inattendu",
        "rest": "Repos",
    }
    rows = []
    for employee in compliance_employees:
        for detail in employee.get("details", []):
            observed = detail.get("observed") or {}
            rows.append(
                {
                    "tenant": employee.get("tenant") or "",
                    "person_id": employee.get("person_id") or "",
                    "employee_name": employee.get("employee_name") or "",
                    "department_name": employee.get("department_name") or "",
                    "planning_name": employee.get("planning_name") or "",
                    "work_shift_name": employee.get("work_shift_name") or "",
                    "date": detail.get("date") or "",
                    "status": status_labels.get(detail.get("status"), detail.get("status") or ""),
                    "expected_work_period": "Oui" if detail.get("expected_work_period") else "Non",
                    "planned_minutes": detail.get("planned_minutes") or 0,
                    "total_logs": observed.get("total_logs") or 0,
                    "checkins": observed.get("checkins") or 0,
                    "checkouts": observed.get("checkouts") or 0,
                    "unknown_events": observed.get("unknown_events") or 0,
                    "arrival_time": _format_export_time(detail.get("actual_checkin_at")),
                    "departure_time": _format_export_time(detail.get("actual_checkout_at")),
                    "expected_checkin_at": _format_export_datetime(detail.get("expected_checkin_at")),
                    "actual_checkin_at": _format_export_datetime(detail.get("actual_checkin_at")),
                    "arrival_delta_minutes": (
                        detail.get("arrival_delta_minutes")
                        if detail.get("arrival_delta_minutes") is not None
                        else ""
                    ),
                    "expected_checkout_at": _format_export_datetime(detail.get("expected_checkout_at")),
                    "actual_checkout_at": _format_export_datetime(detail.get("actual_checkout_at")),
                    "departure_delta_minutes": (
                        detail.get("departure_delta_minutes")
                        if detail.get("departure_delta_minutes") is not None
                        else ""
                    ),
                }
            )
    return rows


def _build_attendance_excel_response(
    *,
    period: str,
    start_date: date,
    end_date: date,
    filters: dict,
    compliance_employees: list[dict],
    selected_field_ids: list[str],
) -> HttpResponse | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return None

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Rapport Pointage"

    worksheet.append(["Rapport de pointage"])
    worksheet.append(["Periode", period, "Debut", start_date.isoformat(), "Fin", end_date.isoformat()])
    worksheet.append(
        [
            "Tenant",
            filters.get("tenant") or "Tous",
            "Departement",
            str(filters.get("department_id") or "Tous"),
            "Personnes",
            ", ".join(filters.get("person_ids") or []) or "Toutes",
        ]
    )
    worksheet.append([])

    detail_rows = _build_attendance_export_rows(compliance_employees)
    headers, export_rows = _build_attendance_export_matrix(detail_rows, selected_field_ids)
    worksheet.append(headers)
    header_row_index = worksheet.max_row
    for col_idx in range(1, len(headers) + 1):
        worksheet.cell(row=header_row_index, column=col_idx).font = Font(bold=True)

    for row_values in export_rows:
        worksheet.append(row_values)

    for column_cells in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = _build_export_filename("attendance-report", start_date, end_date, "xlsx")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_attendance_pdf_response(
    *,
    period: str,
    start_date: date,
    end_date: date,
    filters: dict,
    summary: dict,
    compliance_employees: list[dict],
    selected_field_ids: list[str],
) -> HttpResponse | None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return None

    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Rapport de pointage", styles["Title"]))
    elements.append(Spacer(1, 8))
    elements.append(
        Paragraph(
            f"Periode: {period} | Du {start_date.isoformat()} au {end_date.isoformat()} | "
            f"Tenant: {filters.get('tenant') or 'Tous'}",
            styles["Normal"],
        )
    )
    elements.append(
        Paragraph(
            f"Departement: {filters.get('department_id') or 'Tous'} | "
            f"Personnes: {', '.join(filters.get('person_ids') or []) or 'Toutes'}",
            styles["Normal"],
        )
    )
    elements.append(
        Paragraph(
            f"Logs: {summary.get('total_logs', 0)} | Entrees: {summary.get('checkins', 0)} | "
            f"Sorties: {summary.get('checkouts', 0)} | Inconnus: {summary.get('unknown_events', 0)}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 12))

    table_data = [[
        "Person ID",
        "Employe",
        "Departement",
        "Attendus",
        "Conformes",
        "Partiels",
        "Manquants",
        "Inattendus",
        "Taux %",
    ]]
    for employee in compliance_employees:
        table_data.append(
            [
                employee.get("person_id") or "",
                employee.get("employee_name") or "",
                employee.get("department_name") or "",
                employee.get("expected_work_days") or 0,
                employee.get("compliant_days") or 0,
                employee.get("partial_days") or 0,
                employee.get("missing_days") or 0,
                employee.get("unexpected_activity_days") or 0,
                employee.get("compliance_rate") if employee.get("compliance_rate") is not None else "-",
            ]
        )
    if len(table_data) == 1:
        table_data.append(["-", "Aucune donnee", "-", 0, 0, 0, 0, 0, "-"])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9CA3AF")),
                ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)

    detail_rows = _build_attendance_export_rows(compliance_employees)
    if detail_rows:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Details horaires (arrivee/depart)", styles["Heading3"]))
        detail_headers, detail_matrix_rows = _build_attendance_export_matrix(detail_rows, selected_field_ids)
        details_table_data = [detail_headers]
        for row_values in detail_matrix_rows:
            details_table_data.append(
                [
                    "-" if value is None or value == "" else str(value)
                    for value in row_values
                ]
            )
        if len(details_table_data) == 1:
            details_table_data.append(["-" for _ in detail_headers])
        details_table = Table(details_table_data, repeatRows=1)
        details_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9CA3AF")),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(details_table)
    document.build(elements)

    pdf_bytes = buffer.getvalue()
    buffer.close()
    filename = _build_export_filename("attendance-report", start_date, end_date, "pdf")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_attendance_csv_response(
    *,
    start_date: date,
    end_date: date,
    compliance_employees: list[dict],
    selected_field_ids: list[str],
) -> HttpResponse:
    detail_rows = _build_attendance_export_rows(compliance_employees)
    headers, matrix_rows = _build_attendance_export_matrix(detail_rows, selected_field_ids)

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row_values in matrix_rows:
        writer.writerow(row_values)

    filename = _build_export_filename("attendance-report", start_date, end_date, "csv")
    response = HttpResponse(f"\ufeff{output.getvalue()}", content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def hik_sync_devices_api(request: HttpRequest) -> Response:
    denied = _require_admin_api(request)
    if denied is not None:
        return denied

    lock_key = "hik_gateway:sync-devices-lock"
    if not cache.add(lock_key, "1", timeout=90):
        return Response(
            {"detail": "Device sync already running. Retry in a few seconds."},
            status=status.HTTP_409_CONFLICT,
        )

    dispatch_core_devices = _to_bool(request.data.get("dispatch_core_devices", True))
    try:
        synced = sync_all_gateways()
        dispatched = dispatch_hik_devices_to_core_devices() if dispatch_core_devices else 0
        return Response(
            {
                "status": "ok",
                "synced": synced,
                "dispatched": dispatched,
                "dispatch_core_devices": dispatch_core_devices,
            },
            status=status.HTTP_200_OK,
        )
    finally:
        cache.delete(lock_key)


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def hik_acs_events_api(request: HttpRequest) -> Response:
    payload_source = request.query_params if request.method == "GET" else request.data

    dev_index = str(payload_source.get("dev_index") or "").strip()
    if not dev_index:
        return Response({"detail": "dev_index is required"}, status=status.HTTP_400_BAD_REQUEST)

    payload = payload_source.get("payload")
    if payload is None:
        payload = payload_source
    if not isinstance(payload, dict):
        return Response({"detail": "payload must be an object"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        max_results = int(payload_source.get("max_results") or payload_source.get("maxResults") or 30)
    except (TypeError, ValueError):
        return Response({"detail": "max_results must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

    if max_results <= 0:
        return Response({"detail": "max_results must be > 0"}, status=status.HTTP_400_BAD_REQUEST)

    request_payload = _normalize_acs_event_cond(payload, default_max_results=max_results)
    fallback_payload = _acs_event_payload_without_time_window(request_payload)
    fallback_used = False
    logger.warning(
        "ACS search request received dev_index=%s method=%s payload=%s",
        dev_index,
        request.method,
        _safe_json_preview(request_payload),
    )

    try:
        client = get_shared_gateway_client()
        gateway_response = client.acs_event_search(dev_index, request_payload)
    except Exception as exc:  # noqa: BLE001
        if fallback_payload is None or not _is_bad_json_error_message(exc):
            logger.exception("Unable to fetch AcsEvent from shared gateway")
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
        logger.warning(
            "ACS search fallback triggered dev_index=%s reason=%s fallback_payload=%s",
            dev_index,
            str(exc),
            _safe_json_preview(fallback_payload),
        )
        try:
            gateway_response = client.acs_event_search(dev_index, fallback_payload)
            request_payload = fallback_payload
            fallback_used = True
        except Exception as fallback_exc:  # noqa: BLE001
            logger.exception("Unable to fetch AcsEvent from shared gateway")
            return Response({"detail": str(fallback_exc)}, status=status.HTTP_502_BAD_GATEWAY)

    logger.warning(
        "ACS search response dev_index=%s fallback_used=%s summary=%s",
        dev_index,
        fallback_used,
        _safe_json_preview(_acs_event_summary(gateway_response)),
    )
    if fallback_used:
        tail_payload = _build_tail_acs_event_payload(request_payload, gateway_response)
        if tail_payload is not None:
            logger.warning(
                "ACS search tail-page fetch dev_index=%s payload=%s",
                dev_index,
                _safe_json_preview(tail_payload),
            )
            try:
                tail_response = client.acs_event_search(dev_index, tail_payload)
                request_payload = tail_payload
                gateway_response = tail_response
                logger.warning(
                    "ACS search tail-page response dev_index=%s summary=%s",
                    dev_index,
                    _safe_json_preview(_acs_event_summary(gateway_response)),
                )
            except Exception as tail_exc:  # noqa: BLE001
                logger.warning(
                    "ACS search tail-page fetch failed dev_index=%s detail=%s",
                    dev_index,
                    str(tail_exc),
                )

    return Response(
        {
            "dev_index": dev_index,
            "request": request_payload,
            "response": gateway_response,
        },
        status=status.HTTP_200_OK,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def hik_read_card_api(request: HttpRequest) -> Response:
    dev_index = str(request.data.get("dev_index") or "").strip()
    if not dev_index:
        return Response({"detail": "dev_index is required"}, status=status.HTTP_400_BAD_REQUEST)

    tenant_code = str(request.data.get("tenant") or request.data.get("tenant_code") or "").strip()
    if tenant_code:
        denied, _ = _require_tenant_scope_api(request, tenant_code)
        if denied is not None:
            return denied

    try:
        timeout_seconds = int(request.data.get("timeout_seconds", 15))
    except (TypeError, ValueError):
        return Response({"detail": "timeout_seconds must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        poll_interval_ms = int(request.data.get("poll_interval_ms", 1200))
    except (TypeError, ValueError):
        return Response({"detail": "poll_interval_ms must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

    if timeout_seconds <= 0:
        return Response({"detail": "timeout_seconds must be > 0"}, status=status.HTTP_400_BAD_REQUEST)
    if poll_interval_ms <= 0:
        return Response({"detail": "poll_interval_ms must be > 0"}, status=status.HTTP_400_BAD_REQUEST)

    timeout_seconds = min(timeout_seconds, 60)
    poll_interval_ms = min(poll_interval_ms, 5000)

    if tenant_code:
        device_exists = Device.objects.filter(
            tenant__code__iexact=tenant_code,
            dev_index=dev_index,
        ).exists()
        if not device_exists:
            return Response(
                {"detail": "Aucun lecteur correspondant a ce tenant."},
                status=status.HTTP_404_NOT_FOUND,
            )

    try:
        client = get_shared_gateway_client(tenant_code=tenant_code or None)
    except Exception as exc:  # noqa: BLE001
        return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    started_at = timezone.now().astimezone(dt_timezone.utc)
    deadline = started_at + timedelta(seconds=timeout_seconds)
    search_id = f"card-read-{dev_index}"
    attempt = 0
    baseline_serial: int | None = None
    logger.warning(
        "Read-card session started dev_index=%s tenant=%s timeout_seconds=%s poll_interval_ms=%s started_at=%s",
        dev_index,
        tenant_code or "",
        timeout_seconds,
        poll_interval_ms,
        started_at.isoformat(),
    )

    while timezone.now().astimezone(dt_timezone.utc) <= deadline:
        attempt += 1
        fallback_used = False
        condition_with_window = {
            "AcsEventCond": {
                "searchID": search_id,
                "searchResultPosition": 0,
                "maxResults": 30,
                "startTime": started_at.isoformat(),
                "endTime": timezone.now().astimezone(dt_timezone.utc).isoformat(),
            }
        }
        condition_without_window = {
            "AcsEventCond": {
                "searchID": search_id,
                "searchResultPosition": 0,
                "maxResults": 30,
            }
        }
        try:
            gateway_response = client.acs_event_search(dev_index, condition_with_window)
        except Exception as exc:  # noqa: BLE001
            if not _is_bad_json_error_message(exc):
                logger.exception("Unable to read card from reader", extra={"dev_index": dev_index, "tenant": tenant_code})
                return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
            try:
                gateway_response = client.acs_event_search(dev_index, condition_without_window)
                fallback_used = True
                logger.warning(
                    "Read-card fallback without time window used dev_index=%s tenant=%s attempt=%s",
                    dev_index,
                    tenant_code or "",
                    attempt,
                )
            except Exception as fallback_exc:  # noqa: BLE001
                logger.exception(
                    "Unable to read card from reader (fallback without time window failed)",
                    extra={"dev_index": dev_index, "tenant": tenant_code},
                )
                return Response({"detail": str(fallback_exc)}, status=status.HTTP_502_BAD_GATEWAY)

        if fallback_used:
            tail_payload = _build_tail_acs_event_payload(condition_without_window, gateway_response)
            if tail_payload is not None:
                logger.warning(
                    "Read-card tail-page fetch dev_index=%s tenant=%s attempt=%s payload=%s",
                    dev_index,
                    tenant_code or "",
                    attempt,
                    _safe_json_preview(tail_payload),
                )
                try:
                    gateway_response = client.acs_event_search(dev_index, tail_payload)
                    logger.warning(
                        "Read-card tail-page response dev_index=%s tenant=%s attempt=%s summary=%s",
                        dev_index,
                        tenant_code or "",
                        attempt,
                        _safe_json_preview(_acs_event_summary(gateway_response)),
                    )
                except Exception as tail_exc:  # noqa: BLE001
                    logger.warning(
                        "Read-card tail-page fetch failed dev_index=%s tenant=%s attempt=%s detail=%s",
                        dev_index,
                        tenant_code or "",
                        attempt,
                        str(tail_exc),
                    )

        if baseline_serial is None:
            baseline_serial = _extract_max_event_serial(gateway_response)

        diagnostics: dict = {}
        latest = _extract_latest_card_event(
            gateway_response,
            started_at=started_at,
            min_serial_exclusive=baseline_serial,
            diagnostics=diagnostics,
        )
        logger.warning(
            "Read-card poll result dev_index=%s tenant=%s attempt=%s diagnostics=%s acs_summary=%s",
            dev_index,
            tenant_code or "",
            attempt,
            _safe_json_preview(diagnostics),
            _safe_json_preview(_acs_event_summary(gateway_response)),
        )
        if latest is not None:
            logger.warning(
                "Read-card success dev_index=%s tenant=%s attempt=%s selected=%s",
                dev_index,
                tenant_code or "",
                attempt,
                _safe_json_preview(latest),
            )
            return Response(
                {
                    "status": "ok",
                    "dev_index": dev_index,
                    **latest,
                },
                status=status.HTTP_200_OK,
            )

        pytime.sleep(poll_interval_ms / 1000)

    logger.warning(
        "Read-card timeout dev_index=%s tenant=%s attempts=%s started_at=%s deadline=%s",
        dev_index,
        tenant_code or "",
        attempt,
        started_at.isoformat(),
        deadline.isoformat(),
    )
    return Response(
        {"detail": "Aucune carte detectee pendant la fenetre de lecture."},
        status=status.HTTP_408_REQUEST_TIMEOUT,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def hik_catchup_acs_events_api(request: HttpRequest) -> Response:
    denied = _require_admin_api(request)
    if denied is not None:
        return denied

    try:
        max_results = int(request.data.get("max_results", 50))
    except (TypeError, ValueError):
        return Response({"detail": "max_results must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

    if max_results <= 0:
        return Response({"detail": "max_results must be > 0"}, status=status.HTTP_400_BAD_REQUEST)

    total = catchup_all_devices(max_results=max_results)
    return Response(
        {
            "status": "ok",
            "processed": total,
            "max_results": max_results,
        },
        status=status.HTTP_200_OK,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def hik_register_webhooks_api(request: HttpRequest) -> Response:
    denied = _require_admin_api(request)
    if denied is not None:
        return denied

    ip_address = str(request.data.get("ip_address") or getattr(settings, "HIK_WEBHOOK_IP", "")).strip()
    if not ip_address:
        return Response({"detail": "ip_address is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        port = int(request.data.get("port", getattr(settings, "HIK_WEBHOOK_PORT", 443)))
    except (TypeError, ValueError):
        return Response({"detail": "port must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

    if port <= 0:
        return Response({"detail": "port must be > 0"}, status=status.HTTP_400_BAD_REQUEST)

    url = str(request.data.get("url") or getattr(settings, "HIK_WEBHOOK_URL", "/api/hik/events")).strip()
    if not url:
        return Response({"detail": "url is required"}, status=status.HTTP_400_BAD_REQUEST)

    client = get_shared_gateway_client()
    registered = 0
    for device in Device.objects.all().iterator():
        payload = {
            "HttpHostNotificationList": [
                {
                    "HttpHostNotification": {
                        "id": "1",
                        "url": url,
                        "protocolType": "HTTP",
                        "addressingFormatType": "ipaddress",
                        "ipAddress": ip_address,
                        "portNo": port,
                        "SubscribeEvent": {
                            "heartbeat": 30,
                            "eventMode": "all",
                        },
                    }
                }
            ]
        }
        client.set_http_host(device.dev_index, payload)
        registered += 1

    return Response(
        {
            "status": "ok",
            "registered": registered,
            "ip_address": ip_address,
            "port": port,
            "url": url,
        },
        status=status.HTTP_200_OK,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def hik_events_api(request: HttpRequest) -> Response:
    tenant_code = (request.GET.get("tenant") or "").strip()
    source = (request.GET.get("source") or "").strip().lower()
    dev_index = (request.GET.get("dev_index") or "").strip()
    person_id = (request.GET.get("person_id") or "").strip()
    include_system = _to_bool(request.GET.get("include_system", "0"))
    auto_catchup = _to_bool(request.GET.get("auto_catchup", "1"))
    since_id_raw = (request.GET.get("since_id") or "").strip()
    since_id: int | None = None

    try:
        limit = int(request.GET.get("limit", 100))
    except ValueError:
        return Response({"detail": "limit must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

    if limit <= 0:
        return Response({"detail": "limit must be > 0"}, status=status.HTTP_400_BAD_REQUEST)
    limit = min(limit, 500)

    if since_id_raw:
        try:
            since_id = int(since_id_raw)
        except ValueError:
            return Response({"detail": "since_id must be an integer"}, status=status.HTTP_400_BAD_REQUEST)
        if since_id <= 0:
            return Response({"detail": "since_id must be > 0"}, status=status.HTTP_400_BAD_REQUEST)

    if not tenant_code and not _is_admin_request(request):
        return Response(
            {"detail": "Ajoute ?tenant=<code_tenant> (ou connecte-toi en administrateur pour voir tous les événements)."},
            status=status.HTTP_403_FORBIDDEN,
        )
    if tenant_code:
        denied, _ = _require_tenant_scope_api(request, tenant_code)
        if denied is not None:
            return denied

    try:
        catchup_cooldown_seconds = max(
            2,
            int(getattr(settings, "HIK_EVENTS_AUTO_CATCHUP_THROTTLE_SECONDS", 8)),
        )
    except (TypeError, ValueError):
        catchup_cooldown_seconds = 8

    should_trigger_auto_catchup = False
    if tenant_code and auto_catchup:
        cache_key = f"hik-events-auto-catchup:{tenant_code.lower()}"
        if cache.add(cache_key, "1", timeout=catchup_cooldown_seconds):
            should_trigger_auto_catchup = True

    queryset = AttendanceLog.objects.select_related("device", "tenant", "raw_event").order_by("-timestamp", "-id")
    if tenant_code:
        queryset = queryset.filter(tenant__code__iexact=tenant_code)
    if source:
        queryset = queryset.filter(source__iexact=source)
    if dev_index:
        queryset = queryset.filter(device__dev_index=dev_index)
    if person_id:
        queryset = queryset.filter(person_id=person_id)
    if since_id is not None:
        queryset = queryset.filter(id__gt=since_id)
    if not include_system:
        queryset = queryset.exclude(
            Q(person_id="") & Q(normalized_action=AttendanceLog.ACTION_UNKNOWN)
        )

    logs = list(queryset[:limit])
    tenant_ids = {log.tenant_id for log in logs}
    person_ids = {str(log.person_id).strip() for log in logs if str(log.person_id).strip()}
    employee_lookup: dict[tuple[int, str], Employee] = {}
    if tenant_ids and person_ids:
        employee_qs = Employee.objects.select_related("department").filter(
            tenant_id__in=tenant_ids,
            employee_no__in=person_ids,
        )
        for employee in employee_qs:
            employee_lookup[(employee.tenant_id, employee.employee_no)] = employee

    def _normalize_access_status(log: AttendanceLog) -> str:
        status_text = str(log.attendance_status or "").strip().lower()
        combined = " ".join(
            [
                status_text,
                str(log.attendance_type or "").strip().lower(),
                str(log.direction or "").strip().lower(),
                str(log.raw_event.sub_event_type or "").strip().lower(),
            ]
        )
        denied_tokens = ("deny", "denied", "refus", "failed", "forbid", "unauthor")
        granted_tokens = ("grant", "allow", "autor", "success", "pass", "in")
        if any(token in combined for token in denied_tokens):
            return "denied"
        if any(token in combined for token in granted_tokens):
            return "granted"
        return "unknown"

    results = [
        {
            "id": log.id,
            "tenant": log.tenant.code,
            "timestamp": log.timestamp,
            "person_id": log.person_id,
            "employee_name": (
                employee_lookup[(log.tenant_id, log.person_id)].full_name
                if (log.tenant_id, log.person_id) in employee_lookup
                else ""
            ),
            "department_name": (
                employee_lookup[(log.tenant_id, log.person_id)].department.name
                if (log.tenant_id, log.person_id) in employee_lookup
                and employee_lookup[(log.tenant_id, log.person_id)].department is not None
                else ""
            ),
            "device": {
                "id": log.device_id,
                "dev_index": log.device.dev_index,
                "serial_number": log.device.serial_number,
                "device_name": log.device.device_name,
                "status": log.device.status,
            },
            "attendance_type": log.attendance_type,
            "attendance_status": log.attendance_status,
            "normalized_action": log.normalized_action,
            "access_status": _normalize_access_status(log),
            "direction": log.direction,
            "source": log.source,
            "raw_event": {
                "id": log.raw_event_id,
                "event_type": log.raw_event.event_type,
                "event_datetime": log.raw_event.event_datetime,
                "major_event_type": log.raw_event.major_event_type,
                "sub_event_type": log.raw_event.sub_event_type,
                "serial_no": log.raw_event.serial_no,
                "card_reader_no": log.raw_event.card_reader_no,
                "door_no": log.raw_event.door_no,
            },
        }
        for log in logs
    ]

    payload = {
        "count": len(results),
        "results": results,
        "filters": {
            "tenant": tenant_code or None,
            "source": source or None,
            "dev_index": dev_index or None,
            "person_id": person_id or None,
            "include_system": include_system,
            "since_id": since_id,
            "limit": limit,
        },
    }

    if should_trigger_auto_catchup and tenant_code:
        _start_tenant_auto_catchup(tenant_code=tenant_code, max_results=30)

    return Response(payload)


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def hik_attendance_corrections_api(request: HttpRequest) -> Response:
    tenant_code = str(
        request.query_params.get("tenant")
        if request.method == "GET"
        else request.data.get("tenant")
        or ""
    ).strip()
    person_id = str(
        request.query_params.get("person_id")
        if request.method == "GET"
        else request.data.get("person_id")
        or ""
    ).strip()

    if not tenant_code:
        return Response({"detail": "tenant is required"}, status=status.HTTP_400_BAD_REQUEST)
    if not person_id:
        return Response({"detail": "person_id is required"}, status=status.HTTP_400_BAD_REQUEST)
    denied, tenant = _require_tenant_scope_api(request, tenant_code)
    if denied is not None:
        return denied

    employee = (
        Employee.objects.select_related("tenant")
        .filter(tenant=tenant, employee_no=person_id)
        .first()
    )
    if employee is None:
        return Response({"detail": "employee not found for tenant/person_id"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        date_value = str(request.query_params.get("date") or "").strip()
        start_date_value = str(request.query_params.get("start_date") or "").strip()
        end_date_value = str(request.query_params.get("end_date") or "").strip()

        if date_value:
            start_date_value = date_value
            end_date_value = date_value

        try:
            start_date = _parse_iso_date(start_date_value, "start_date") if start_date_value else None
            end_date = _parse_iso_date(end_date_value, "end_date") if end_date_value else None
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        queryset = AttendanceCorrection.objects.select_related("tenant", "employee", "created_by", "updated_by").filter(
            tenant=employee.tenant,
            employee=employee,
        )
        if start_date is not None:
            queryset = queryset.filter(work_date__gte=start_date)
        if end_date is not None:
            queryset = queryset.filter(work_date__lte=end_date)
        if start_date and end_date and end_date < start_date:
            return Response({"detail": "end_date must be greater than or equal to start_date"}, status=status.HTTP_400_BAD_REQUEST)

        corrections = list(queryset.order_by("work_date", "id"))
        return Response({"count": len(corrections), "results": [_serialize_attendance_correction(item) for item in corrections]})

    date_value = str(request.data.get("date") or "").strip()
    arrival_key_present = "arrival_time" in request.data
    departure_key_present = "departure_time" in request.data
    break_start_key_present = "break_start_time" in request.data
    break_end_key_present = "break_end_time" in request.data
    overtime_key_present = "overtime_hours" in request.data
    notes_key_present = "notes" in request.data
    try:
        work_date = _parse_iso_date(date_value, "date")
        arrival_time = _parse_iso_time(request.data.get("arrival_time"), "arrival_time")
        departure_time = _parse_iso_time(request.data.get("departure_time"), "departure_time")
        break_start_time = _parse_iso_time(request.data.get("break_start_time"), "break_start_time")
        break_end_time = _parse_iso_time(request.data.get("break_end_time"), "break_end_time")
        overtime_hours = _parse_optional_decimal(request.data.get("overtime_hours"), "overtime_hours")
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    if work_date is None:
        return Response({"detail": "date is required"}, status=status.HTTP_400_BAD_REQUEST)

    if bool(break_start_time) != bool(break_end_time):
        return Response(
            {"detail": "break_start_time and break_end_time must be provided together"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if break_start_time and break_end_time and break_start_time == break_end_time:
        return Response({"detail": "break_end_time must be different from break_start_time"}, status=status.HTTP_400_BAD_REQUEST)
    if arrival_time and departure_time and arrival_time == departure_time:
        return Response({"detail": "departure_time must be different from arrival_time"}, status=status.HTTP_400_BAD_REQUEST)

    notes = str(request.data.get("notes") or "").strip()
    user = request.user if getattr(request, "user", None) and request.user.is_authenticated else None

    previous_correction = AttendanceCorrection.objects.select_related("tenant", "employee", "created_by", "updated_by").filter(
        tenant=employee.tenant,
        employee=employee,
        work_date=work_date,
    ).first()
    previous_payload = _serialize_attendance_correction(previous_correction) if previous_correction is not None else None

    if previous_correction is None:
        provided_any_value = any(
            [
                arrival_key_present and arrival_time is not None,
                departure_key_present and departure_time is not None,
                break_start_key_present and break_start_time is not None,
                break_end_key_present and break_end_time is not None,
                overtime_key_present and overtime_hours is not None,
                notes_key_present and bool(notes),
            ]
        )
        if not provided_any_value:
            return Response({"detail": "at least one correction field is required"}, status=status.HTTP_400_BAD_REQUEST)
        correction = AttendanceCorrection(
            tenant=employee.tenant,
            employee=employee,
            work_date=work_date,
            arrival_time=arrival_time if arrival_key_present else None,
            departure_time=departure_time if departure_key_present else None,
            break_start_time=break_start_time if break_start_key_present else None,
            break_end_time=break_end_time if break_end_key_present else None,
            overtime_hours=overtime_hours if overtime_key_present else None,
            notes=notes if notes_key_present else "",
        )
        if user is not None:
            correction.created_by = user
            correction.updated_by = user
        correction.save()
        created = True
    else:
        correction = previous_correction
        if arrival_key_present:
            correction.arrival_time = arrival_time
        if departure_key_present:
            correction.departure_time = departure_time
        if break_start_key_present or break_end_key_present:
            correction.break_start_time = break_start_time
            correction.break_end_time = break_end_time
        if overtime_key_present:
            correction.overtime_hours = overtime_hours
        if notes_key_present:
            correction.notes = notes
        if user is not None:
            correction.updated_by = user
        correction.save()
        created = False

    if (
        correction.arrival_time is None
        and correction.departure_time is None
        and correction.break_start_time is None
        and correction.break_end_time is None
        and correction.overtime_hours is None
        and not str(correction.notes or "").strip()
    ):
        correction.delete()
        return Response(
            {
                "status": "deleted",
                "result": None,
            },
            status=status.HTTP_200_OK,
        )

    if created and user is not None:
        correction.created_by = user
        correction.updated_by = user
        correction.save(update_fields=["created_by", "updated_by", "updated_at"])

    correction = AttendanceCorrection.objects.select_related("tenant", "employee", "created_by", "updated_by").get(
        id=correction.id
    )
    current_payload = _serialize_attendance_correction(correction)
    AttendanceCorrectionLog.objects.create(
        correction=correction,
        tenant=employee.tenant,
        employee=employee,
        work_date=work_date,
        action=AttendanceCorrectionLog.ACTION_CREATE if created else AttendanceCorrectionLog.ACTION_UPDATE,
        payload={
            "before": previous_payload,
            "after": current_payload,
        },
        changed_by=user,
    )

    return Response(
        {
            "status": "created" if created else "updated",
            "result": current_payload,
        },
        status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def hik_attendance_correction_logs_api(request: HttpRequest) -> Response:
    tenant_code = str(request.query_params.get("tenant") or "").strip()
    person_id = str(request.query_params.get("person_id") or "").strip()
    date_value = str(request.query_params.get("date") or "").strip()
    start_date_value = str(request.query_params.get("start_date") or "").strip()
    end_date_value = str(request.query_params.get("end_date") or "").strip()

    if not tenant_code:
        return Response({"detail": "tenant is required"}, status=status.HTTP_400_BAD_REQUEST)
    denied, tenant = _require_tenant_scope_api(request, tenant_code)
    if denied is not None:
        return denied

    if date_value:
        start_date_value = date_value
        end_date_value = date_value

    try:
        start_date = _parse_iso_date(start_date_value, "start_date") if start_date_value else None
        end_date = _parse_iso_date(end_date_value, "end_date") if end_date_value else None
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    queryset = AttendanceCorrectionLog.objects.select_related(
        "tenant",
        "employee",
        "changed_by",
        "correction",
    ).filter(tenant=tenant)
    if person_id:
        queryset = queryset.filter(employee__employee_no=person_id)
    if start_date is not None:
        queryset = queryset.filter(work_date__gte=start_date)
    if end_date is not None:
        queryset = queryset.filter(work_date__lte=end_date)
    if start_date and end_date and end_date < start_date:
        return Response({"detail": "end_date must be greater than or equal to start_date"}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(queryset.order_by("-created_at", "-id")[:500])
    return Response({"count": len(rows), "results": [_serialize_attendance_correction_log(item) for item in rows]})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def hik_attendance_reports_api(request: HttpRequest) -> Response:
    tenant_code = (request.GET.get("tenant") or "").strip()
    period = (request.GET.get("period") or "daily").strip().lower()
    export_format = (request.GET.get("export") or "json").strip().lower()
    person_id = (request.GET.get("person_id") or "").strip()
    auto_catchup = _to_bool(request.GET.get("auto_catchup", "1"))
    person_ids = _parse_csv_query_list(request.GET.get("person_ids", ""))
    if person_id:
        person_ids.append(person_id)
    person_ids = sorted({value for value in person_ids if value})
    dev_index = (request.GET.get("dev_index") or "").strip()
    site = (request.GET.get("site") or "").strip()
    manager = (request.GET.get("manager") or "").strip()
    source = (request.GET.get("source") or "").strip().lower()
    department_id_value = (request.GET.get("department_id") or "").strip()
    anomaly_type = (request.GET.get("anomaly_type") or "").strip().lower()
    anomaly_types = sorted(set(_parse_csv_query_list(anomaly_type))) if anomaly_type else []
    validation_status = (request.GET.get("validation_status") or "").strip().lower()
    validation_statuses = sorted(set(_parse_csv_query_list(validation_status))) if validation_status else []
    requested_export_fields = _parse_csv_query_list(request.GET.get("fields", ""))
    selected_export_fields = _resolve_attendance_export_fields(requested_export_fields)

    if period not in {"daily", "weekly", "monthly"}:
        return Response(
            {"detail": "period must be one of: daily, weekly, monthly"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if export_format not in {"json", "excel", "xlsx", "pdf", "csv"}:
        return Response(
            {"detail": "format must be one of: json, excel, xlsx, pdf, csv"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    department_id = None
    if department_id_value:
        try:
            department_id = int(department_id_value)
        except (TypeError, ValueError):
            return Response({"detail": "department_id must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

    if not tenant_code and not _is_admin_request(request):
        return Response(
            {"detail": "Ajoute ?tenant=<code_tenant> (ou connecte-toi en administrateur pour voir tous les rapports)."},
            status=status.HTTP_403_FORBIDDEN,
        )
    if tenant_code:
        denied, _ = _require_tenant_scope_api(request, tenant_code)
        if denied is not None:
            return denied

    try:
        start_date, end_date = _resolve_report_window(
            period,
            start_date_value=request.GET.get("start_date"),
            end_date_value=request.GET.get("end_date"),
            reference_date_value=request.GET.get("date"),
        )
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    current_tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.combine(start_date, time.min), current_tz)
    end_dt = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), time.min), current_tz)
    query_start_dt = start_dt - timedelta(days=1)
    query_end_dt = end_dt + timedelta(days=1)

    queryset = AttendanceLog.objects.select_related(
        "device",
        "tenant",
        "raw_event",
        "employee",
        "employee__department",
    ).filter(
        timestamp__gte=query_start_dt,
        timestamp__lt=query_end_dt,
    ).order_by("timestamp", "id")

    if tenant_code:
        queryset = queryset.filter(tenant__code__iexact=tenant_code)
    if source:
        queryset = queryset.filter(source__iexact=source)
    if dev_index:
        queryset = queryset.filter(device__dev_index=dev_index)
    if person_ids:
        queryset = queryset.filter(person_id__in=person_ids)
    if department_id is not None:
        queryset = queryset.filter(employee__department_id=department_id)
    if site:
        queryset = queryset.filter(
            Q(device__device_name__icontains=site)
            | Q(device__dev_index__iexact=site)
        )
    if manager:
        queryset = queryset.filter(
            employee__attributes__name__iexact="manager",
            employee__attributes__value__icontains=manager,
        )
    if manager or site:
        queryset = queryset.distinct()

    logs = list(queryset)
    resolver = ScheduleResolver()
    timeline_map: dict[str, dict] = {}
    employee_map: dict[tuple[int, str], dict] = {}
    employee_day_observed: dict[tuple[int, str, str], dict] = {}
    summary = {
        "total_logs": 0,
        "total_employees": 0,
        "checkins": 0,
        "checkouts": 0,
        "unknown_events": 0,
    }

    for log in logs:
        local_timestamp = timezone.localtime(log.timestamp, current_tz)
        target_date = local_timestamp.date()
        direction = _classify_attendance_direction(log)
        matched_shift_payload = None
        if log.employee_id is not None:
            matched_shift = resolver.resolve_shift_from_timestamp(
                log.employee,
                log.timestamp,
                direction_hint=direction,
            )
            if matched_shift is not None:
                target_date = matched_shift.shift_date
                matched_shift_payload = {
                    "id": matched_shift.work_shift.id,
                    "name": matched_shift.work_shift.name,
                    "code": matched_shift.work_shift.code,
                    "start_time": matched_shift.work_shift.start_time,
                    "end_time": matched_shift.work_shift.end_time,
                }

        if target_date < start_date or target_date > end_date:
            continue

        date_key = target_date.isoformat()
        employee_key = (log.tenant_id, str(log.person_id or "").strip() or f"unresolved:{log.id}")
        employee_name = log.employee.full_name if log.employee_id else ""
        department_name = (
            log.employee.department.name
            if log.employee_id and log.employee.department_id is not None
            else ""
        )

        day_bucket = timeline_map.setdefault(
            date_key,
            {
                "date": date_key,
                "total_logs": 0,
                "distinct_employee_keys": set(),
                "checkins": 0,
                "checkouts": 0,
                "unknown_events": 0,
            },
        )
        employee_bucket = employee_map.setdefault(
            employee_key,
            {
                "tenant": log.tenant.code,
                "person_id": str(log.person_id or "").strip(),
                "employee_name": employee_name,
                "department_name": department_name,
                "total_logs": 0,
                "checkins": 0,
                "checkouts": 0,
                "unknown_events": 0,
                "days_present": set(),
                "first_activity": local_timestamp,
                "last_activity": local_timestamp,
                "first_checkin": None,
                "last_checkout": None,
            },
        )

        summary["total_logs"] += 1
        day_bucket["total_logs"] += 1
        day_bucket["distinct_employee_keys"].add(employee_key)
        employee_bucket["total_logs"] += 1
        employee_bucket["days_present"].add(date_key)
        employee_bucket["first_activity"] = min(employee_bucket["first_activity"], local_timestamp)
        employee_bucket["last_activity"] = max(employee_bucket["last_activity"], local_timestamp)
        observed_bucket = employee_day_observed.setdefault(
            (log.tenant_id, str(log.person_id or "").strip(), date_key),
            {
                "total_logs": 0,
                "checkins": 0,
                "checkouts": 0,
                "unknown_events": 0,
                "first_checkin": None,
                "last_checkout": None,
                "first_activity": None,
                "last_activity": None,
                "matched_shift": None,
            },
        )
        observed_bucket["total_logs"] += 1
        if observed_bucket["first_activity"] is None or local_timestamp < observed_bucket["first_activity"]:
            observed_bucket["first_activity"] = local_timestamp
        if observed_bucket["last_activity"] is None or local_timestamp > observed_bucket["last_activity"]:
            observed_bucket["last_activity"] = local_timestamp
        if matched_shift_payload is not None and observed_bucket["matched_shift"] is None:
            observed_bucket["matched_shift"] = matched_shift_payload

        if direction == "IN":
            summary["checkins"] += 1
            day_bucket["checkins"] += 1
            employee_bucket["checkins"] += 1
            observed_bucket["checkins"] += 1
            if employee_bucket["first_checkin"] is None or local_timestamp < employee_bucket["first_checkin"]:
                employee_bucket["first_checkin"] = local_timestamp
            if observed_bucket["first_checkin"] is None or local_timestamp < observed_bucket["first_checkin"]:
                observed_bucket["first_checkin"] = local_timestamp
        elif direction == "OUT":
            summary["checkouts"] += 1
            day_bucket["checkouts"] += 1
            employee_bucket["checkouts"] += 1
            observed_bucket["checkouts"] += 1
            if employee_bucket["last_checkout"] is None or local_timestamp > employee_bucket["last_checkout"]:
                employee_bucket["last_checkout"] = local_timestamp
            if observed_bucket["last_checkout"] is None or local_timestamp > observed_bucket["last_checkout"]:
                observed_bucket["last_checkout"] = local_timestamp
        else:
            summary["unknown_events"] += 1
            day_bucket["unknown_events"] += 1
            employee_bucket["unknown_events"] += 1
            observed_bucket["unknown_events"] += 1

    summary["total_employees"] = len(employee_map)
    timeline = []
    for day in sorted(timeline_map):
        bucket = timeline_map[day]
        timeline.append(
            {
                "date": bucket["date"],
                "total_logs": bucket["total_logs"],
                "distinct_employees": len(bucket["distinct_employee_keys"]),
                "checkins": bucket["checkins"],
                "checkouts": bucket["checkouts"],
                "unknown_events": bucket["unknown_events"],
            }
        )

    employees = []
    for _, bucket in sorted(
        employee_map.items(),
        key=lambda item: (
            item[1]["employee_name"] or item[1]["person_id"] or "",
            item[1]["tenant"],
        ),
    ):
        employees.append(
            {
                "tenant": bucket["tenant"],
                "person_id": bucket["person_id"],
                "employee_name": bucket["employee_name"],
                "department_name": bucket["department_name"],
                "total_logs": bucket["total_logs"],
                "checkins": bucket["checkins"],
                "checkouts": bucket["checkouts"],
                "unknown_events": bucket["unknown_events"],
                "days_present": len(bucket["days_present"]),
                "first_activity": bucket["first_activity"],
                "last_activity": bucket["last_activity"],
                "first_checkin": bucket["first_checkin"],
                "last_checkout": bucket["last_checkout"],
            }
        )

    employee_scope_qs = Employee.objects.select_related(
        "tenant",
        "department",
    ).filter(
        is_active=True,
    )
    if tenant_code:
        employee_scope_qs = employee_scope_qs.filter(tenant__code__iexact=tenant_code)
    if department_id is not None:
        employee_scope_qs = employee_scope_qs.filter(department_id=department_id)
    if person_ids:
        employee_scope_qs = employee_scope_qs.filter(employee_no__in=person_ids)
    if manager:
        employee_scope_qs = employee_scope_qs.filter(
            attributes__name__iexact="manager",
            attributes__value__icontains=manager,
        )
    if manager:
        employee_scope_qs = employee_scope_qs.distinct()

    observed_employee_keys = {
        (log.tenant_id, str(log.person_id or "").strip())
        for log in logs
        if str(log.person_id or "").strip()
    }
    scoped_by_event_filters = bool(dev_index or source or site or anomaly_types or validation_statuses)
    if scoped_by_event_filters and not person_ids and department_id is None and not observed_employee_keys:
        employee_scope_qs = employee_scope_qs.none()
    if not person_ids and department_id is None and observed_employee_keys:
        tenant_ids = {key[0] for key in observed_employee_keys}
        employee_nos = {key[1] for key in observed_employee_keys}
        employee_scope_qs = employee_scope_qs.filter(tenant_id__in=tenant_ids, employee_no__in=employee_nos)

    corrections_qs = AttendanceCorrection.objects.select_related("tenant", "employee", "created_by", "updated_by").filter(
        work_date__gte=start_date,
        work_date__lte=end_date,
    )
    if tenant_code:
        corrections_qs = corrections_qs.filter(tenant__code__iexact=tenant_code)
    if department_id is not None:
        corrections_qs = corrections_qs.filter(employee__department_id=department_id)
    if person_ids:
        corrections_qs = corrections_qs.filter(employee__employee_no__in=person_ids)
    elif observed_employee_keys:
        correction_tenant_ids = {key[0] for key in observed_employee_keys}
        correction_employee_nos = {key[1] for key in observed_employee_keys}
        corrections_qs = corrections_qs.filter(
            tenant_id__in=correction_tenant_ids,
            employee__employee_no__in=correction_employee_nos,
        )
    elif scoped_by_event_filters:
        corrections_qs = corrections_qs.none()
    if manager:
        corrections_qs = corrections_qs.filter(
            employee__attributes__name__iexact="manager",
            employee__attributes__value__icontains=manager,
        )
    if manager:
        corrections_qs = corrections_qs.distinct()

    corrections = list(corrections_qs.order_by("work_date", "id"))
    correction_by_key = {
        (item.tenant_id, item.employee.employee_no, item.work_date.isoformat()): item
        for item in corrections
    }
    correction_logs_qs = AttendanceCorrectionLog.objects.select_related(
        "tenant",
        "employee",
        "changed_by",
        "correction",
    ).filter(
        work_date__gte=start_date,
        work_date__lte=end_date,
    )
    if tenant_code:
        correction_logs_qs = correction_logs_qs.filter(tenant__code__iexact=tenant_code)
    if department_id is not None:
        correction_logs_qs = correction_logs_qs.filter(employee__department_id=department_id)
    if person_ids:
        correction_logs_qs = correction_logs_qs.filter(employee__employee_no__in=person_ids)
    elif observed_employee_keys:
        correction_tenant_ids = {key[0] for key in observed_employee_keys}
        correction_employee_nos = {key[1] for key in observed_employee_keys}
        correction_logs_qs = correction_logs_qs.filter(
            tenant_id__in=correction_tenant_ids,
            employee__employee_no__in=correction_employee_nos,
        )
    elif scoped_by_event_filters:
        correction_logs_qs = correction_logs_qs.none()
    if manager:
        correction_logs_qs = correction_logs_qs.filter(
            employee__attributes__name__iexact="manager",
            employee__attributes__value__icontains=manager,
        )
    if manager:
        correction_logs_qs = correction_logs_qs.distinct()
    correction_logs = list(correction_logs_qs.order_by("created_at", "id"))
    correction_log_by_key = {}
    for item in correction_logs:
        correction_log_by_key[(item.tenant_id, item.employee.employee_no, item.work_date.isoformat())] = item

    compliance_summary = {
        "evaluated_employees": 0,
        "expected_work_days": 0,
        "compliant_days": 0,
        "partial_days": 0,
        "missing_days": 0,
        "unexpected_activity_days": 0,
        "rest_days": 0,
        "compliance_rate": None,
    }
    compliance_employees = []
    total_expected_minutes = 0
    total_worked_minutes = 0
    total_normal_minutes = 0
    total_overtime_minutes = 0
    punctuality_total = 0
    punctuality_on_time = 0

    for employee in employee_scope_qs.order_by("name", "employee_no", "id"):
        if not person_ids and department_id is None and observed_employee_keys:
            employee_key = (employee.tenant_id, employee.employee_no)
            if employee_key not in observed_employee_keys:
                continue

        expected_work_days = 0
        compliant_days = 0
        partial_days = 0
        missing_days = 0
        unexpected_activity_days = 0
        rest_days = 0
        day_details = []

        resolved_planning = resolver.resolve_effective_planning(employee, start_date)
        resolved_shift = resolver.resolve_effective_work_shift(employee, start_date)

        for target_day in _iterate_dates(start_date, end_date):
            date_key = target_day.isoformat()
            observed = employee_day_observed.get(
                (employee.tenant_id, employee.employee_no, date_key),
                {
                    "total_logs": 0,
                    "checkins": 0,
                    "checkouts": 0,
                    "unknown_events": 0,
                    "first_checkin": None,
                    "last_checkout": None,
                    "first_activity": None,
                    "last_activity": None,
                },
            )
            correction = correction_by_key.get((employee.tenant_id, employee.employee_no, date_key))
            correction_log = correction_log_by_key.get((employee.tenant_id, employee.employee_no, date_key))
            day_schedule = resolver.build_day_schedule(employee, target_day)
            has_work_period = bool(day_schedule.get("has_work_period")) and not bool(day_schedule.get("is_rest_day"))
            has_checkin = observed["checkins"] > 0
            has_checkout = observed["checkouts"] > 0
            expected_checkin_at = None
            expected_checkout_at = None
            if has_work_period:
                expected_checkin_at, expected_checkout_at = _expected_bounds_from_day_schedule(
                    day_schedule,
                    target_day=target_day,
                    current_tz=current_tz,
                )
            actual_checkin_at = observed.get("first_checkin")
            actual_checkout_at = observed.get("last_checkout")
            first_activity_at = observed.get("first_activity")
            last_activity_at = observed.get("last_activity")

            # Some sites use a single reader (no explicit OUT event). In reports,
            # use first/last observed access as arrival/departure fallback.
            if actual_checkin_at is None and first_activity_at is not None:
                has_checkin = True
                actual_checkin_at = first_activity_at
            if actual_checkout_at is None and last_activity_at is not None:
                if actual_checkin_at is not None:
                    if last_activity_at > actual_checkin_at:
                        has_checkout = True
                        actual_checkout_at = last_activity_at
                elif int(observed.get("total_logs") or 0) >= 2:
                    has_checkout = True
                    actual_checkout_at = last_activity_at

            if correction is not None:
                if correction.arrival_time is not None:
                    has_checkin = True
                    actual_checkin_at = timezone.make_aware(datetime.combine(target_day, correction.arrival_time), current_tz)
                if correction.departure_time is not None:
                    has_checkout = True
                    departure_day = target_day
                    if correction.arrival_time is not None and correction.departure_time <= correction.arrival_time:
                        departure_day = target_day + timedelta(days=1)
                    actual_checkout_at = timezone.make_aware(
                        datetime.combine(departure_day, correction.departure_time),
                        current_tz,
                    )
            matched_shift = observed.get("matched_shift")
            arrival_delta_minutes = _minutes_delta(actual_checkin_at, expected_checkin_at)
            departure_delta_minutes = _minutes_delta(actual_checkout_at, expected_checkout_at)
            worked_minutes = _calculate_worked_minutes(actual_checkin_at, actual_checkout_at)
            planned_minutes = int(day_schedule.get("planned_minutes", 0) or 0)
            overtime_enabled = _is_overtime_enabled_for_day(
                resolver=resolver,
                employee=employee,
                target_day=target_day,
            )
            overtime_minutes = max(worked_minutes - planned_minutes, 0)
            if not overtime_enabled:
                overtime_minutes = 0
            if correction is not None and correction.overtime_hours is not None:
                overtime_minutes = (
                    max(int(Decimal(correction.overtime_hours) * Decimal(60)), 0)
                    if overtime_enabled
                    else 0
                )
            normal_minutes = max(min(worked_minutes, planned_minutes), 0)
            covers_at_least_one_shift = _covers_at_least_one_shift(
                day_schedule=day_schedule,
                target_day=target_day,
                current_tz=current_tz,
                actual_checkin_at=actual_checkin_at,
                actual_checkout_at=actual_checkout_at,
            )
            validation_state = _extract_validation_status(correction_log.payload if correction_log is not None else None)
            if not validation_state and correction is not None:
                validation_state = "validated" if correction.updated_by_id or correction.created_by_id else "to_review"
            if not validation_state:
                validation_state = "none"
            anomalies = _build_detail_anomalies(
                status_key="",
                has_work_period=has_work_period,
                has_checkin=has_checkin,
                has_checkout=has_checkout,
                arrival_delta_minutes=arrival_delta_minutes,
                departure_delta_minutes=departure_delta_minutes,
                unknown_events=observed.get("unknown_events", 0),
            )

            if has_work_period:
                expected_work_days += 1
                total_expected_minutes += planned_minutes
                if has_checkin and has_checkout:
                    if covers_at_least_one_shift:
                        status_key = "compliant"
                        compliant_days += 1
                    else:
                        status_key = "partial"
                        partial_days += 1
                elif has_checkin or has_checkout:
                    status_key = "partial"
                    partial_days += 1
                else:
                    status_key = "missing"
                    missing_days += 1
            else:
                if observed["total_logs"] > 0 or has_checkin or has_checkout:
                    status_key = "unexpected_activity"
                    unexpected_activity_days += 1
                else:
                    status_key = "rest"
                    rest_days += 1

            anomalies = _build_detail_anomalies(
                status_key=status_key,
                has_work_period=has_work_period,
                has_checkin=has_checkin,
                has_checkout=has_checkout,
                arrival_delta_minutes=arrival_delta_minutes,
                departure_delta_minutes=departure_delta_minutes,
                unknown_events=observed.get("unknown_events", 0),
            )
            if has_work_period:
                punctuality_total += 1
                if arrival_delta_minutes is not None and arrival_delta_minutes <= 0:
                    punctuality_on_time += 1
            total_worked_minutes += worked_minutes
            total_normal_minutes += normal_minutes
            total_overtime_minutes += overtime_minutes

            hr_status = "conforme"
            if any(code in anomalies for code in {"absence", "pointage_incomplet", "activite_inattendue"}):
                hr_status = "anomalie"
            elif anomalies:
                hr_status = "a_verifier"

            if anomaly_types and not any(code in anomaly_types for code in anomalies):
                continue
            if validation_statuses and validation_state not in validation_statuses:
                continue

            if status_key != "rest":
                day_details.append(
                    {
                        "date": date_key,
                        "status": status_key,
                        "hr_status": hr_status,
                        "validation_status": validation_state,
                        "anomaly_types": anomalies,
                        "expected_work_period": has_work_period,
                        "observed": observed,
                        "planned_minutes": planned_minutes,
                        "expected_checkin_at": expected_checkin_at,
                        "expected_checkout_at": expected_checkout_at,
                        "actual_checkin_at": actual_checkin_at,
                        "actual_checkout_at": actual_checkout_at,
                        "arrival_delta_minutes": arrival_delta_minutes,
                        "departure_delta_minutes": departure_delta_minutes,
                        "worked_minutes": worked_minutes,
                        "worked_hours": _minutes_to_hours(worked_minutes),
                        "normal_minutes": normal_minutes,
                        "normal_hours": _minutes_to_hours(normal_minutes),
                        "overtime_minutes": overtime_minutes,
                        "overtime_hours": _minutes_to_hours(overtime_minutes),
                        "late_minutes": max(arrival_delta_minutes or 0, 0),
                        "early_leave_minutes": abs(min(departure_delta_minutes or 0, 0)),
                        "covers_at_least_one_shift": covers_at_least_one_shift,
                        "overtime_enabled": overtime_enabled,
                        "matched_shift": matched_shift,
                        "correction": _serialize_attendance_correction(correction) if correction is not None else None,
                        "correction_log": _serialize_attendance_correction_log(correction_log) if correction_log is not None else None,
                    }
                )

        if (anomaly_types or validation_statuses) and not day_details:
            continue

        employee_compliance_rate = (
            round((compliant_days / expected_work_days) * 100, 2)
            if expected_work_days > 0
            else None
        )
        late_days = sum(1 for detail in day_details if (detail.get("late_minutes") or 0) > 0)
        early_leave_days = sum(1 for detail in day_details if (detail.get("early_leave_minutes") or 0) > 0)
        anomaly_days = sum(1 for detail in day_details if (detail.get("hr_status") or "") == "anomalie")
        total_worked_employee_minutes = sum(int(detail.get("worked_minutes") or 0) for detail in day_details)
        total_normal_employee_minutes = sum(int(detail.get("normal_minutes") or 0) for detail in day_details)
        total_overtime_employee_minutes = sum(int(detail.get("overtime_minutes") or 0) for detail in day_details)
        compliance_employees.append(
            {
                "tenant": employee.tenant.code,
                "person_id": employee.employee_no,
                "matricule": employee.employee_no,
                "employee_name": employee.full_name or employee.employee_no,
                "department_name": employee.department.name if employee.department_id else "",
                "service": employee.department.name if employee.department_id else "",
                "position": employee.position or "",
                "planning_name": resolved_planning.name if resolved_planning else "",
                "work_shift_name": resolved_shift.name if resolved_shift else "",
                "expected_work_days": expected_work_days,
                "compliant_days": compliant_days,
                "partial_days": partial_days,
                "missing_days": missing_days,
                "unexpected_activity_days": unexpected_activity_days,
                "rest_days": rest_days,
                "late_days": late_days,
                "early_leave_days": early_leave_days,
                "anomaly_days": anomaly_days,
                "total_worked_minutes": total_worked_employee_minutes,
                "total_worked_hours": _minutes_to_hours(total_worked_employee_minutes),
                "total_normal_minutes": total_normal_employee_minutes,
                "total_normal_hours": _minutes_to_hours(total_normal_employee_minutes),
                "total_overtime_minutes": total_overtime_employee_minutes,
                "total_overtime_hours": _minutes_to_hours(total_overtime_employee_minutes),
                "compliance_rate": employee_compliance_rate,
                "details": day_details,
            }
        )

        compliance_summary["evaluated_employees"] += 1
        compliance_summary["expected_work_days"] += expected_work_days
        compliance_summary["compliant_days"] += compliant_days
        compliance_summary["partial_days"] += partial_days
        compliance_summary["missing_days"] += missing_days
        compliance_summary["unexpected_activity_days"] += unexpected_activity_days
        compliance_summary["rest_days"] += rest_days

    if compliance_summary["expected_work_days"] > 0:
        compliance_summary["compliance_rate"] = round(
            (compliance_summary["compliant_days"] / compliance_summary["expected_work_days"]) * 100,
            2,
        )
    employee_total = len(compliance_employees)
    present_employees = sum(1 for item in compliance_employees if int(item.get("compliant_days", 0)) + int(item.get("partial_days", 0)) > 0)
    absent_employees = sum(1 for item in compliance_employees if int(item.get("missing_days", 0)) > 0 and int(item.get("compliant_days", 0)) + int(item.get("partial_days", 0)) == 0)
    late_employees = sum(1 for item in compliance_employees if int(item.get("late_days", 0)) > 0)
    repeated_absences = [
        {
            "person_id": item.get("person_id"),
            "employee_name": item.get("employee_name"),
            "missing_days": item.get("missing_days"),
        }
        for item in compliance_employees
        if int(item.get("missing_days", 0)) >= 2
    ][:5]
    recurring_anomalies = [
        {
            "person_id": item.get("person_id"),
            "employee_name": item.get("employee_name"),
            "anomaly_days": item.get("anomaly_days"),
        }
        for item in compliance_employees
        if int(item.get("anomaly_days", 0)) >= 2
    ][:5]
    executive_summary = {
        "effectif_total": employee_total,
        "attendance_overview": {
            "present_employees": present_employees,
            "absent_employees": absent_employees,
            "late_employees": late_employees,
            "present_days": compliance_summary["compliant_days"] + compliance_summary["partial_days"],
            "absent_days": compliance_summary["missing_days"],
        },
        "heures": {
            "normal_minutes": total_normal_minutes,
            "normal_hours": _minutes_to_hours(total_normal_minutes),
            "overtime_minutes": total_overtime_minutes,
            "overtime_hours": _minutes_to_hours(total_overtime_minutes),
            "worked_minutes": total_worked_minutes,
            "worked_hours": _minutes_to_hours(total_worked_minutes),
            "planned_minutes": total_expected_minutes,
            "planned_hours": _minutes_to_hours(total_expected_minutes),
        },
        "taux_ponctualite_global": (
            round((punctuality_on_time / punctuality_total) * 100, 2)
            if punctuality_total > 0
            else None
        ),
        "alertes_majeures": {
            "absences_repetees": repeated_absences,
            "anomalies_recurrentes": recurring_anomalies,
        },
    }

    filters_payload = {
        "tenant": tenant_code or None,
        "person_id": person_id or None,
        "person_ids": person_ids,
        "department_id": department_id,
        "dev_index": dev_index or None,
        "site": site or None,
        "manager": manager or None,
        "source": source or None,
        "anomaly_type": anomaly_types,
        "validation_status": validation_statuses,
        "export_fields": selected_export_fields,
    }
    response_payload = {
        "period": period,
        "range": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
        },
        "summary": summary,
        "executive_summary": executive_summary,
        "timeline": timeline,
        "employees": employees,
        "filters": filters_payload,
        "corrections": [_serialize_attendance_correction(item) for item in corrections],
        "correction_history": [_serialize_attendance_correction_log(item) for item in correction_logs],
        "compliance": {
            "summary": compliance_summary,
            "employees": compliance_employees,
        },
    }

    if export_format in {"excel", "xlsx"}:
        export_response = _build_attendance_excel_response(
            period=period,
            start_date=start_date,
            end_date=end_date,
            filters=filters_payload,
            compliance_employees=compliance_employees,
            selected_field_ids=selected_export_fields,
        )
        if export_response is None:
            return Response(
                {"detail": "Excel export requires openpyxl. Install dependency and retry."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        return export_response

    if export_format == "pdf":
        export_response = _build_attendance_pdf_response(
            period=period,
            start_date=start_date,
            end_date=end_date,
            filters=filters_payload,
            summary=summary,
            compliance_employees=compliance_employees,
            selected_field_ids=selected_export_fields,
        )
        if export_response is None:
            return Response(
                {"detail": "PDF export requires reportlab. Install dependency and retry."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        return export_response

    if export_format == "csv":
        return _build_attendance_csv_response(
            start_date=start_date,
            end_date=end_date,
            compliance_employees=compliance_employees,
            selected_field_ids=selected_export_fields,
        )

    return Response(response_payload, status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def hik_devices_api(request: HttpRequest) -> Response:
    tenant_code = (request.GET.get("tenant") or "").strip()
    unassigned_only = _to_bool(request.GET.get("unassigned_only", "0"))
    protocol_query = (request.GET.get("protocol") or "").strip()
    status_query = (request.GET.get("status") or "").strip()
    dev_type = (request.GET.get("dev_type") or "").strip()
    key = (request.GET.get("key") or "").strip()
    normalized = _to_bool(request.GET.get("normalized", "1"))

    try:
        max_result = int(request.GET.get("max_result", 100))
    except ValueError:
        return Response({"detail": "max_result must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

    protocol_types = _parse_csv_query_list(protocol_query)
    statuses = _parse_csv_query_list(status_query)

    if not tenant_code and not _is_admin_request(request):
        if unassigned_only and _is_authenticated_request(request):
            pass
        else:
            return Response(
                {"detail": "Ajoute ?tenant=<code_tenant> (ou connecte-toi en administrateur pour voir tous les appareils)."},
                status=status.HTTP_403_FORBIDDEN,
            )

    if tenant_code and unassigned_only:
        return Response(
            {"detail": "Utilise soit ?tenant=<code_tenant>, soit ?unassigned_only=1, mais pas les deux."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if tenant_code:
        denied, _ = _require_tenant_scope_api(request, tenant_code)
        if denied is not None:
            return denied

    devices = []
    errors = []

    tenant_by_dev_index = {
        dev_index: tenant__code
        for dev_index, tenant__code in Device.objects.select_related("tenant").values_list("dev_index", "tenant__code")
    }
    allowed_dev_indexes = None
    if tenant_code:
        allowed_dev_indexes = {
            dev_index
            for dev_index, mapped_tenant_code in tenant_by_dev_index.items()
            if str(mapped_tenant_code).lower() == tenant_code.lower()
        }
        if not allowed_dev_indexes:
            allowed_dev_indexes = None

    try:
        client = get_shared_gateway_client(tenant_code=tenant_code or None)
        payload = client.device_list_all(
            max_result=max_result,
            protocol_types=protocol_types or None,
            statuses=statuses or None,
            dev_type=dev_type,
            key=key,
        )
    except Exception as exc:  # noqa: BLE001
        logger.exception("Unable to list devices from shared gateway")
        errors.append(str(exc))
        payload = {}

    gateway_payloads = [{"tenant_code": tenant_code or "*", "search_result": payload.get("SearchResult", {})}]

    for item in extract_devices(payload):
        normalized_item = normalize_device(item)
        dev_index_value = normalized_item.get("dev_index") or item.get("devIndex", "")
        if allowed_dev_indexes is not None and dev_index_value not in allowed_dev_indexes:
            continue

        normalized_item["sn"] = (item.get("EhomeParams", {}) or {}).get("EhomeID", "")
        normalized_item["devIndex"] = item.get("devIndex", "")
        normalized_item["name"] = item.get("devName") or item.get("deviceName") or ""
        normalized_item["model"] = item.get("devType") or item.get("deviceType") or ""
        normalized_item["version"] = item.get("devVersion") or ""
        normalized_item["dev_serial"] = item.get("devSerial") or item.get("serialNumber") or ""
        normalized_item["offline_hint"] = item.get("offlineHint") or item.get("offlineReason") or ""
        normalized_item["tenant_code"] = tenant_by_dev_index.get(dev_index_value, "")
        normalized_item["gateway_base_url"] = "shared"
        if unassigned_only and normalized_item["tenant_code"]:
            continue
        devices.append(normalized_item)

    if not normalized:
        return Response({"count": len(gateway_payloads), "results": gateway_payloads, "errors": errors})

    return Response({"count": len(devices), "results": devices, "errors": errors})


@require_GET
def hik_devices_page(request: HttpRequest):
    tenant_code = (request.GET.get("tenant") or "").strip()
    unassigned_only = _to_bool(request.GET.get("unassigned_only", "0"))
    is_admin = _is_admin_request(request)

    request_parameters = request.GET.get("request", "").strip() or json.dumps(
        DEFAULT_DEVICE_LIST_PAYLOAD,
        ensure_ascii=False,
        indent=2,
    )
    response_format = (request.GET.get("format") or "").strip().lower()
    wants_json = response_format == "json" or "application/json" in request.headers.get("Accept", "")

    context = {
        "devices": [],
        "tenant_code": tenant_code,
        "error": "",
        "is_admin": is_admin,
        "request_parameters": request_parameters,
        "response_parameters": "",
        "status_code": "-",
        "gateway_url": "",
    }

    if not tenant_code and not is_admin:
        if not (unassigned_only and _is_authenticated_request(request)):
            error_message = "Ajoute ?tenant=<code_tenant> (ou connecte-toi en administrateur pour voir tous les appareils)."
            if wants_json:
                return JsonResponse({"detail": error_message}, status=403)
            context["error"] = error_message
            return render(request, "hik_gateway/device_list.html", context, status=403)

    if tenant_code and unassigned_only:
        error_message = "Utilise soit ?tenant=<code_tenant>, soit ?unassigned_only=1, mais pas les deux."
        if wants_json:
            return JsonResponse({"detail": error_message}, status=400)
        context["error"] = error_message
        return render(request, "hik_gateway/device_list.html", context, status=400)
    if tenant_code and _is_authenticated_request(request) and not is_admin:
        denied, _ = _require_tenant_scope_api(request, tenant_code)
        if denied is not None:
            if wants_json:
                return JsonResponse(denied.data, status=denied.status_code)
            context["error"] = denied.data.get("detail", "Tenant access denied.")
            return render(request, "hik_gateway/device_list.html", context, status=denied.status_code)


    devices = []
    errors = []
    response_payload: dict | None = None

    try:
        payload_to_send = json.loads(request_parameters)
    except json.JSONDecodeError:
        error_message = "Request Parameters doit être un JSON valide."
        if wants_json:
            return JsonResponse({"detail": error_message}, status=400)
        context["error"] = error_message
        return render(request, "hik_gateway/device_list.html", context, status=400)

    tenant_by_dev_index = {
        dev_index: tenant__code
        for dev_index, tenant__code in Device.objects.select_related("tenant").values_list("dev_index", "tenant__code")
    }
    allowed_dev_indexes = None
    if tenant_code:
        allowed_dev_indexes = {
            dev_index
            for dev_index, mapped_tenant_code in tenant_by_dev_index.items()
            if str(mapped_tenant_code).lower() == tenant_code.lower()
        }
        if not allowed_dev_indexes:
            allowed_dev_indexes = None

    try:
        client = get_shared_gateway_client(tenant_code=tenant_code or None)
        context["gateway_url"] = "shared"
        payload = client.device_list(payload=payload_to_send)
        response_payload = payload
        context["status_code"] = 200
    except Exception as exc:  # noqa: BLE001
        errors.append(str(exc))
        payload = {}

    for item in extract_devices(payload):
        normalized = normalize_device(item)
        dev_index_value = normalized.get("dev_index") or item.get("devIndex", "")
        if allowed_dev_indexes is not None and dev_index_value not in allowed_dev_indexes:
            continue
        normalized["tenant_code"] = tenant_by_dev_index.get(dev_index_value, "")
        if unassigned_only and normalized["tenant_code"]:
            continue
        normalized["gateway_base_url"] = "shared"
        devices.append(normalized)

    if response_payload is not None:
        context["response_parameters"] = json.dumps(response_payload, ensure_ascii=False, indent=2)

    if errors and not devices:
        context["error"] = "Impossible de récupérer les devices: " + " | ".join(errors)
    elif errors:
        context["error"] = "La connexion gateway a échoué: " + " | ".join(errors)

    context["devices"] = devices
    if wants_json:
        return JsonResponse(
            {
                "count": len(devices),
                "results": devices,
                "errors": errors,
                "tenant": tenant_code or None,
                "unassigned_only": unassigned_only,
                "status_code": context["status_code"],
                "response_parameters": response_payload,
            }
        )

    return render(request, "hik_gateway/device_list.html", context)


@require_GET
def hikdevice_devices_space(request: HttpRequest):
    """Interface simple pour visualiser les appareils disponibles sur HikDevice."""
    tenant_code = (request.GET.get("tenant") or "").strip()
    unassigned_only = _to_bool(request.GET.get("unassigned_only", "0"))
    protocol_query = (request.GET.get("protocol") or "").strip()
    status_query = (request.GET.get("status") or "").strip()
    dev_type = (request.GET.get("dev_type") or "").strip()
    key = (request.GET.get("key") or "").strip()
    is_admin = _is_admin_request(request)

    protocol_types = _parse_csv_query_list(protocol_query)
    statuses = _parse_csv_query_list(status_query)

    context = {
        "devices": [],
        "tenant_code": tenant_code,
        "protocol": protocol_query,
        "status_filter": status_query,
        "dev_type": dev_type,
        "key": key,
        "error": "",
        "is_admin": is_admin,
    }

    if not tenant_code and not is_admin:
        if not (unassigned_only and _is_authenticated_request(request)):
            context["error"] = "Ajoute ?tenant=<code_tenant> (ou connecte-toi en administrateur pour voir tous les appareils)."
            return render(request, "hik_gateway/hikdevice_devices_space.html", context, status=403)

    if tenant_code and unassigned_only:
        context["error"] = "Utilise soit ?tenant=<code_tenant>, soit ?unassigned_only=1, mais pas les deux."
        return render(request, "hik_gateway/hikdevice_devices_space.html", context, status=400)
    if tenant_code and _is_authenticated_request(request) and not is_admin:
        denied, _ = _require_tenant_scope_api(request, tenant_code)
        if denied is not None:
            context["error"] = denied.data.get("detail", "Tenant access denied.")
            return render(
                request,
                "hik_gateway/hikdevice_devices_space.html",
                context,
                status=denied.status_code,
            )

    tenant_by_dev_index = {
        dev_index: tenant__code
        for dev_index, tenant__code in Device.objects.select_related("tenant").values_list("dev_index", "tenant__code")
    }

    allowed_dev_indexes = None
    if tenant_code:
        allowed_dev_indexes = {
            dev_index
            for dev_index, mapped_tenant_code in tenant_by_dev_index.items()
            if str(mapped_tenant_code).lower() == tenant_code.lower()
        }
        if not allowed_dev_indexes:
            allowed_dev_indexes = None

    try:
        client = get_shared_gateway_client(tenant_code=tenant_code or None)
        payload = client.device_list_all(
            max_result=100,
            protocol_types=protocol_types or None,
            statuses=statuses or None,
            dev_type=dev_type,
            key=key,
        )
    except Exception as exc:  # noqa: BLE001
        logger.exception("Unable to list devices in hikdevice devices space")
        context["error"] = f"Impossible de récupérer les appareils: {exc}"
        return render(request, "hik_gateway/hikdevice_devices_space.html", context, status=502)

    devices = []
    for item in extract_devices(payload):
        normalized_item = normalize_device(item)
        dev_index_value = normalized_item.get("dev_index") or item.get("devIndex", "")
        if allowed_dev_indexes is not None and dev_index_value not in allowed_dev_indexes:
            continue

        normalized_item["sn"] = (item.get("EhomeParams", {}) or {}).get("EhomeID", "")
        normalized_item["tenant_code"] = tenant_by_dev_index.get(dev_index_value, "")
        if unassigned_only and normalized_item["tenant_code"]:
            continue
        devices.append(normalized_item)

    context["devices"] = devices
    return render(request, "hik_gateway/hikdevice_devices_space.html", context)
